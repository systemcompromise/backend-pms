import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MitraStatusDashboardExporter:
    def __init__(self):
        self.primary_color = "1E3A8A"
        self.secondary_color = "3B82F6"
        self.success_color = "10B981"
        self.warning_color = "F59E0B"
        self.danger_color = "EF4444"
        self.light_bg = "F3F4F6"
        self.header_bg = "1E40AF"
        
    def create_workbook_with_data(self, data, output_path):
        wb = openpyxl.Workbook()
        
        self.create_executive_summary(wb, data)
        self.create_status_distribution_sheet(wb, data)
        self.create_monthly_trends_sheet(wb, data)
        self.create_weekly_trends_sheet(wb, data)
        self.create_rider_metrics_sheet(wb, data)
        self.create_visual_charts(wb, data)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.active = wb['Executive Summary']
        wb.save(output_path)
        return output_path
    
    def create_executive_summary(self, wb, data):
        ws = wb.create_sheet("Executive Summary", 0)
        
        summary = data.get('summary', {})
        rider_metrics = data.get('riderMetrics', {})
        filters = data.get('appliedFilters', {})
        
        title = ws.cell(row=2, column=2, value="MITRA LIFECYCLE DASHBOARD - EXECUTIVE SUMMARY")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:G2")
        ws.row_dimensions[2].height = 35
        
        subtitle = ws.cell(row=3, column=2, value="Comprehensive Partner Journey Analytics")
        subtitle.font = Font(size=12, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("B3:G3")
        
        ws.cell(row=5, column=2, value="REPORT INFORMATION").font = Font(bold=True, size=12, color=self.primary_color)
        
        ws.cell(row=7, column=2, value="Generated:").font = Font(bold=True, size=10)
        ws.cell(row=7, column=3, value=datetime.now().strftime("%d %B %Y, %H:%M"))
        
        ws.cell(row=8, column=2, value="Filter Year:").font = Font(bold=True, size=10)
        ws.cell(row=8, column=3, value=filters.get('year') or 'All Years')
        
        ws.cell(row=9, column=2, value="Filter Month:").font = Font(bold=True, size=10)
        ws.cell(row=9, column=3, value=filters.get('month') or 'All Months')
        
        ws.cell(row=10, column=2, value="Filter Week:").font = Font(bold=True, size=10)
        ws.cell(row=10, column=3, value=filters.get('week') or 'All Weeks')
        
        ws.cell(row=12, column=2, value="KEY METRICS").font = Font(bold=True, size=12, color=self.primary_color)
        
        metrics = [
            ("Total Partners", summary.get('totalMitras', 0), "partners", "E3F2FD"),
            ("Active Riders", rider_metrics.get('currentActiveRiders', 0), "riders", "E8F5E9"),
            ("In Training", summary.get('trainingCount', 0), "partners", "FFF3E0"),
            ("Pending Verification", summary.get('pendingCount', 0), "partners", "F3E5F5")
        ]
        
        metric_row = 14
        for label, value, unit, bg_color in metrics:
            ws.cell(row=metric_row, column=2, value=label).font = Font(bold=True, size=10)
            ws.cell(row=metric_row, column=2).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            ws.cell(row=metric_row, column=3, value=value).font = Font(size=12, bold=True, color=self.secondary_color)
            ws.cell(row=metric_row, column=3).number_format = '#,##0'
            ws.cell(row=metric_row, column=3).alignment = Alignment(horizontal="right")
            
            ws.cell(row=metric_row, column=4, value=unit).font = Font(size=9, italic=True)
            metric_row += 1
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def create_status_distribution_sheet(self, wb, data):
        ws = wb.create_sheet("Status Distribution")
        
        status_dist = data.get('statusDistribution', [])
        
        title = ws.cell(row=1, column=1, value="MITRA STATUS DISTRIBUTION")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:E1")
        
        headers = ["Status", "Count", "Percentage", "Category", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        sorted_status = sorted(status_dist, key=lambda x: x.get('count', 0), reverse=True)
        
        for idx, item in enumerate(sorted_status, 4):
            ws.cell(row=idx, column=1, value=item.get('status', 'Unknown')).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=item.get('count', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=3, value=item.get('percentage', 0)).number_format = '0.00"%"'
            
            status = item.get('status', '')
            if status == 'Active':
                category = "Operational"
                color = "D1FAE5"
            elif status in ['Driver Training', 'New', 'Registered']:
                category = "Onboarding"
                color = "FEF3C7"
            else:
                category = "Inactive"
                color = "FEE2E2"
            
            ws.cell(row=idx, column=4, value=category)
            ws.cell(row=idx, column=4).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            ws.cell(row=idx, column=5, value="Primary status" if idx == 4 else "")
        
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_monthly_trends_sheet(self, wb, data):
        ws = wb.create_sheet("Monthly Trends")
        
        monthly_data = data.get('monthlyData', [])
        
        title = ws.cell(row=1, column=1, value="MONTHLY LIFECYCLE TRENDS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:N1")
        
        headers = [
            "Month", "Year", "Active Riders", "Inactive Riders", "Active Status", 
            "New", "Training", "Registered", "Total", "New Joining", 
            "Retention %", "Churn %", "Growth Rate", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for idx, item in enumerate(monthly_data, 4):
            ws.cell(row=idx, column=1, value=item.get('month', ''))
            ws.cell(row=idx, column=2, value=item.get('year', ''))
            ws.cell(row=idx, column=3, value=item.get('riderActiveCount', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=4, value=item.get('riderInactiveCount', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=5, value=item.get('statusCounts', {}).get('Active', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=6, value=item.get('statusCounts', {}).get('New', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=7, value=item.get('statusCounts', {}).get('Driver Training', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=8, value=item.get('statusCounts', {}).get('Registered', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=9, value=item.get('total', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=10, value=item.get('gettingValue', 0)).number_format = '#,##0'
            
            retention = item.get('retentionRate')
            if retention is not None:
                ws.cell(row=idx, column=11, value=retention).number_format = '0.00"%"'
            else:
                ws.cell(row=idx, column=11, value='-')
            
            churn = item.get('churnRate')
            if churn is not None:
                ws.cell(row=idx, column=12, value=churn).number_format = '0.00"%"'
            else:
                ws.cell(row=idx, column=12, value='-')
            
            if idx > 4:
                prev_total = monthly_data[idx-5].get('total', 0)
                curr_total = item.get('total', 0)
                if prev_total > 0:
                    growth = ((curr_total - prev_total) / prev_total) * 100
                    ws.cell(row=idx, column=13, value=growth).number_format = '+0.00%;-0.00%'
                else:
                    ws.cell(row=idx, column=13, value=0).number_format = '0.00%'
            else:
                ws.cell(row=idx, column=13, value='-')
            
            active_count = item.get('riderActiveCount', 0)
            if active_count > 50:
                ws.cell(row=idx, column=14, value="High Activity")
            elif active_count > 20:
                ws.cell(row=idx, column=14, value="Moderate")
            else:
                ws.cell(row=idx, column=14, value="Low Activity")
        
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_weekly_trends_sheet(self, wb, data):
        ws = wb.create_sheet("Weekly Trends")
        
        weekly_data = data.get('weeklyData', [])
        
        title = ws.cell(row=1, column=1, value="WEEKLY LIFECYCLE TRENDS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:O1")
        
        headers = [
            "Week", "Month", "Year", "Active Riders", "Inactive Riders", 
            "Active Status", "New", "Training", "Registered", "Total", 
            "New Joining", "Retention %", "Churn %", "Week Status", "Performance"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for idx, item in enumerate(weekly_data, 4):
            ws.cell(row=idx, column=1, value=item.get('week', ''))
            ws.cell(row=idx, column=2, value=item.get('month', ''))
            ws.cell(row=idx, column=3, value=item.get('year', ''))
            ws.cell(row=idx, column=4, value=item.get('activeCount', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=5, value=item.get('inactiveCount', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=6, value=item.get('statusCounts', {}).get('Active', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=7, value=item.get('statusCounts', {}).get('New', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=8, value=item.get('statusCounts', {}).get('Driver Training', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=9, value=item.get('statusCounts', {}).get('Registered', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=10, value=item.get('total', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=11, value=item.get('gettingValue', 0)).number_format = '#,##0'
            
            retention = item.get('retentionRate')
            if retention is not None:
                ws.cell(row=idx, column=12, value=retention).number_format = '0.00"%"'
            else:
                ws.cell(row=idx, column=12, value='-')
            
            churn = item.get('churnRate')
            if churn is not None:
                ws.cell(row=idx, column=13, value=churn).number_format = '0.00"%"'
            else:
                ws.cell(row=idx, column=13, value='-')
            
            active = item.get('activeCount', 0)
            ws.cell(row=idx, column=14, value="Active Week" if active > 0 else "Inactive")
            
            if retention and retention > 80:
                ws.cell(row=idx, column=15, value="Excellent")
            elif retention and retention > 60:
                ws.cell(row=idx, column=15, value="Good")
            else:
                ws.cell(row=idx, column=15, value="Needs Improvement")
        
        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_rider_metrics_sheet(self, wb, data):
        ws = wb.create_sheet("Rider Metrics")
        
        rider_metrics = data.get('riderMetrics', {})
        
        title = ws.cell(row=1, column=1, value="RIDER PERFORMANCE METRICS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:D1")
        
        ws.cell(row=3, column=1, value="MONTHLY METRICS").font = Font(bold=True, size=12, color=self.secondary_color)
        
        monthly_metrics = [
            ("Current Active Riders", rider_metrics.get('currentActiveRiders', 0)),
            ("Current Inactive Riders", rider_metrics.get('currentInactiveRiders', 0))
        ]
        
        row = 5
        for label, value in monthly_metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).number_format = '#,##0'
            ws.cell(row=row, column=2).font = Font(size=12, bold=True, color=self.secondary_color)
            row += 1
        
        ws.cell(row=row + 1, column=1, value="WEEKLY METRICS").font = Font(bold=True, size=12, color=self.secondary_color)
        
        weekly_metrics = [
            ("Current Week Active Riders", rider_metrics.get('currentWeekActiveRiders', 0)),
            ("Current Week Inactive Riders", rider_metrics.get('currentWeekInactiveRiders', 0))
        ]
        
        row += 3
        for label, value in weekly_metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).number_format = '#,##0'
            ws.cell(row=row, column=2).font = Font(size=12, bold=True, color=self.secondary_color)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def create_visual_charts(self, wb, data):
        ws = wb.create_sheet("Visual Charts")
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value="MITRA LIFECYCLE VISUAL ANALYTICS")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        status_dist = data.get('statusDistribution', [])
        if status_dist:
            chart_ws = wb.create_sheet("Status Chart Data")
            chart_ws.sheet_state = 'hidden'
            
            chart_ws.cell(row=1, column=1, value="Status")
            chart_ws.cell(row=1, column=2, value="Count")
            
            for idx, item in enumerate(status_dist, 2):
                chart_ws.cell(row=idx, column=1, value=item.get('status', ''))
                chart_ws.cell(row=idx, column=2, value=item.get('count', 0))
            
            pie_chart = PieChart()
            pie_chart.title = "Mitra Status Distribution"
            pie_chart.height = 15
            pie_chart.width = 20
            
            data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=1+len(status_dist))
            cats_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=1+len(status_dist))
            
            pie_chart.add_data(data_ref, titles_from_data=True)
            pie_chart.set_categories(cats_ref)
            
            ws.add_chart(pie_chart, "A3")
        
        monthly_data = data.get('monthlyData', [])
        if monthly_data and len(monthly_data) > 1:
            trend_ws = wb.create_sheet("Monthly Trend Data")
            trend_ws.sheet_state = 'hidden'
            
            trend_ws.cell(row=1, column=1, value="Period")
            trend_ws.cell(row=1, column=2, value="Active Riders")
            trend_ws.cell(row=1, column=3, value="Total Partners")
            
            for idx, item in enumerate(monthly_data, 2):
                period = f"{item.get('month', '')} {item.get('year', '')}"
                trend_ws.cell(row=idx, column=1, value=period)
                trend_ws.cell(row=idx, column=2, value=item.get('riderActiveCount', 0))
                trend_ws.cell(row=idx, column=3, value=item.get('total', 0))
            
            line_chart = LineChart()
            line_chart.title = "Monthly Activity Trends"
            line_chart.style = 12
            line_chart.y_axis.title = "Count"
            line_chart.x_axis.title = "Period"
            line_chart.height = 15
            line_chart.width = 28
            
            data_ref = Reference(trend_ws, min_col=2, max_col=3, min_row=1, max_row=1+len(monthly_data))
            cats_ref = Reference(trend_ws, min_col=1, min_row=2, max_row=1+len(monthly_data))
            
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            
            ws.add_chart(line_chart, "A30")
        
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 3

def main():
    try:
        if len(sys.argv) != 3:
            raise ValueError("Usage: python mitraStatusDashboardExporter.py <input_json> <output_excel>")
        
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        exporter = MitraStatusDashboardExporter()
        result_path = exporter.create_workbook_with_data(data, output_path)
        
        print(json.dumps({
            "success": True,
            "output_path": result_path,
            "message": "Mitra status dashboard exported successfully"
        }))
    
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()