import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class TaskAnalyticsChartGenerator:
    def __init__(self):
        self.primary_color = "1E3A8A"
        self.secondary_color = "3B82F6"
        self.success_color = "10B981"
        self.warning_color = "F59E0B"
        self.danger_color = "EF4444"
        self.purple_color = "8B5CF6"
        self.light_bg = "F3F4F6"
        self.header_bg = "1E40AF"
        
    def create_workbook_with_charts(self, data, output_path):
        wb = openpyxl.Workbook()
        self.create_cover_sheet(wb, data)
        self.create_executive_summary(wb, data)
        self.create_performance_sheet(wb, data)
        self.create_insights_sheet(wb, data)
        self.create_recommendations_sheet(wb, data)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_path)
        return output_path
    
    def create_cover_sheet(self, wb, data):
        ws = wb.create_sheet("Dashboard Overview", 0)
        
        title_cell = ws.cell(row=2, column=2, value="TASK MANAGEMENT PERFORMANCE ANALYTICS")
        title_cell.font = Font(bold=True, size=24, color=self.primary_color)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:H2")
        
        subtitle_cell = ws.cell(row=3, column=2, value="Comprehensive Task Performance & User Analytics Report")
        subtitle_cell.font = Font(size=14, color="6B7280", italic=True)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B3:H3")
        
        ws.cell(row=5, column=2, value="Report Generated:").font = Font(bold=True, size=11)
        ws.cell(row=5, column=3, value=datetime.now().strftime("%d %B %Y, %H:%M")).font = Font(size=11)
        
        date_range = data.get('dateRange')
        if date_range:
            ws.cell(row=6, column=2, value="Report Period:").font = Font(bold=True, size=11)
            ws.cell(row=6, column=3, value=f"{date_range['start']} - {date_range['end']}").font = Font(size=11)
        else:
            ws.cell(row=6, column=2, value="Report Period:").font = Font(bold=True, size=11)
            ws.cell(row=6, column=3, value="All Time Data").font = Font(size=11)
        
        ws.cell(row=7, column=2, value="Department:").font = Font(bold=True, size=11)
        ws.cell(row=7, column=3, value="Task Management & Data Analytics").font = Font(size=11)
        
        summary_data = data.get('summaryData', [])
        if summary_data:
            ws.cell(row=9, column=2, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=14, color=self.primary_color)
            
            metric_row = 11
            for item in summary_data[:4]:
                metric_name = ws.cell(row=metric_row, column=2, value=item.get('Metric', ''))
                metric_name.font = Font(bold=True, size=11)
                metric_name.fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
                
                metric_value = ws.cell(row=metric_row, column=3, value=item.get('Value', ''))
                metric_value.font = Font(size=13, bold=True, color=self.secondary_color)
                metric_value.alignment = Alignment(horizontal="right")
                
                metric_unit = ws.cell(row=metric_row, column=4, value=item.get('Unit', ''))
                metric_unit.font = Font(size=10, color="6B7280")
                
                metric_row += 2
        
        ws.cell(row=20, column=2, value="REPORT SECTIONS").font = Font(bold=True, size=12, color=self.primary_color)
        
        sections = [
            ("Executive Summary", "High-level KPIs and performance metrics"),
            ("User Performance Analysis", "Detailed breakdown by user with rankings"),
            ("Strategic Insights", "Top performers and priority improvement areas"),
            ("Management Recommendations", "Data-driven action items for leadership")
        ]
        
        section_row = 22
        for section, desc in sections:
            ws.cell(row=section_row, column=2, value=f"• {section}").font = Font(bold=True, size=11)
            ws.cell(row=section_row, column=3, value=desc).font = Font(size=10, color="6B7280", italic=True)
            section_row += 1
        
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        ws.row_dimensions[2].height = 35
        ws.row_dimensions[3].height = 25
    
    def create_executive_summary(self, wb, data):
        ws = wb.create_sheet("Executive Summary")
        summary_data = data.get('summaryData', [])
        
        if not summary_data:
            return
        
        title = ws.cell(row=1, column=1, value="EXECUTIVE SUMMARY - KEY METRICS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = Font(size=10, color="6B7280")
        
        headers = ["Metric", "Value", "Unit", "Status", "Performance Indicator", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        metrics_map = {}
        for idx, item in enumerate(summary_data, 5):
            metric_name = item.get('Metric', '')
            metrics_map[metric_name] = idx
            
            metric_cell = ws.cell(row=idx, column=1, value=metric_name)
            metric_cell.font = Font(bold=True, size=11)
            metric_cell.fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            metric_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell = ws.cell(row=idx, column=2, value=item.get('Value', 0))
            value_cell.font = Font(size=12, bold=True)
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            unit_cell = ws.cell(row=idx, column=3, value=item.get('Unit', ''))
            unit_cell.font = Font(size=10, color="6B7280")
            unit_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if item.get('Unit') == 'percentage':
                value_cell.value = None
                value_cell.number_format = '0.0%'
            
            desc_cell = ws.cell(row=idx, column=6, value=item.get('Description', ''))
            desc_cell.font = Font(size=10, color="374151")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        total_tasks_row = metrics_map.get('Total Tasks')
        total_eligible_row = metrics_map.get('Total Eligible')
        total_not_eligible_row = metrics_map.get('Total Not Eligible')
        active_users_row = metrics_map.get('Active Users')
        total_invited_row = metrics_map.get('Total Invited')
        total_changed_mind_row = metrics_map.get('Total Changed Mind')
        total_no_response_row = metrics_map.get('Total No Response')
        
        if total_tasks_row and total_eligible_row:
            success_rate_row = total_tasks_row + 1
            ws.insert_rows(success_rate_row)
            
            ws.cell(row=success_rate_row, column=1, value="Overall Success Rate").font = Font(bold=True, size=11)
            ws.cell(row=success_rate_row, column=1).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            
            success_formula = f"=IF(B{total_tasks_row}>0,B{total_eligible_row}/B{total_tasks_row},0)"
            success_cell = ws.cell(row=success_rate_row, column=2)
            success_cell.value = success_formula
            success_cell.number_format = '0.0%'
            success_cell.font = Font(size=12, bold=True)
            success_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            ws.cell(row=success_rate_row, column=3, value="percentage").font = Font(size=10, color="6B7280")
            ws.cell(row=success_rate_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=success_rate_row, column=6, value="Percentage of tasks marked as Eligible").font = Font(size=10, color="374151")
            ws.cell(row=success_rate_row, column=6).alignment = Alignment(wrap_text=True, vertical="center")
            
            metrics_map = {k: (v + 1 if v >= success_rate_row else v) for k, v in metrics_map.items()}
            
            total_eligible_row += 1
            total_not_eligible_row += 1
            active_users_row += 1
            total_invited_row += 1
            total_changed_mind_row += 1
            total_no_response_row += 1
        
        if total_tasks_row and active_users_row:
            avg_tasks_row = active_users_row + 1
            ws.insert_rows(avg_tasks_row)
            
            ws.cell(row=avg_tasks_row, column=1, value="Average Tasks per User").font = Font(bold=True, size=11)
            ws.cell(row=avg_tasks_row, column=1).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            
            avg_formula = f"=IF(B{active_users_row}>0,B{total_tasks_row}/B{active_users_row},0)"
            avg_cell = ws.cell(row=avg_tasks_row, column=2)
            avg_cell.value = avg_formula
            avg_cell.number_format = '0.0'
            avg_cell.font = Font(size=12, bold=True)
            avg_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            ws.cell(row=avg_tasks_row, column=3, value="tasks").font = Font(size=10, color="6B7280")
            ws.cell(row=avg_tasks_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=avg_tasks_row, column=6, value="Average workload distribution across all users").font = Font(size=10, color="374151")
            ws.cell(row=avg_tasks_row, column=6).alignment = Alignment(wrap_text=True, vertical="center")
            
            total_invited_row += 1
            total_changed_mind_row += 1
            total_no_response_row += 1
        
        if total_tasks_row and total_invited_row and total_changed_mind_row and total_no_response_row:
            response_rate_row = total_no_response_row + 1
            ws.insert_rows(response_rate_row)
            
            ws.cell(row=response_rate_row, column=1, value="Response Rate").font = Font(bold=True, size=11)
            ws.cell(row=response_rate_row, column=1).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            
            response_formula = f"=IF(B{total_tasks_row}>0,(B{total_invited_row}+B{total_changed_mind_row}+B{total_no_response_row})/B{total_tasks_row},0)"
            response_cell = ws.cell(row=response_rate_row, column=2)
            response_cell.value = response_formula
            response_cell.number_format = '0.0%'
            response_cell.font = Font(size=12, bold=True)
            response_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            ws.cell(row=response_rate_row, column=3, value="percentage").font = Font(size=10, color="6B7280")
            ws.cell(row=response_rate_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=response_rate_row, column=6, value="Percentage of tasks with responses").font = Font(size=10, color="374151")
            ws.cell(row=response_rate_row, column=6).alignment = Alignment(wrap_text=True, vertical="center")
        
        for row_idx in range(4, ws.max_row + 1):
            value_cell = ws.cell(row=row_idx, column=2)
            if value_cell.value and isinstance(value_cell.value, str) and value_cell.value.startswith('='):
                status = "Excellent"
                status_color = self.success_color
                indicator = "↑"
                
                ws.cell(row=row_idx, column=4, value=status).font = Font(bold=True, color="FFFFFF", size=10)
                ws.cell(row=row_idx, column=4).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
                
                ws.cell(row=row_idx, column=5, value=indicator).font = Font(bold=True, size=14, color=status_color)
                ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws.cell(row=row_idx, column=4, value="N/A").font = Font(bold=True, color="FFFFFF", size=10)
                ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")
                ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
                
                ws.cell(row=row_idx, column=5, value="-").font = Font(bold=True, size=14, color="6B7280")
                ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
        
        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 28
        
        ws.freeze_panes = "A5"
    
    def create_performance_sheet(self, wb, data):
        ws = wb.create_sheet("User Performance Analysis")
        performance_data = data.get('performanceData', [])
        
        if not performance_data:
            return
        
        sorted_performance = sorted(performance_data, key=lambda x: int(x.get('Rank', 999)))
        
        title = ws.cell(row=1, column=1, value="USER PERFORMANCE ANALYSIS & RANKINGS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:N1")
        
        ws.cell(row=2, column=1, value=f"Total Users: {len(sorted_performance)}").font = Font(size=10, color="6B7280")
        
        headers = ["Rank", "User Name", "Total Tasks", "Eligible", "Not Eligible", "Success Rate %", 
                   "Invited", "Changed Mind", "No Response", "Response Rate %", "Conversion Rate %", 
                   "Performance Level", "Projects", "Cities"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for idx, item in enumerate(sorted_performance, 1):
            row = idx + 4
            
            rank_cell = ws.cell(row=row, column=1, value=item.get('Rank', idx))
            rank_cell.alignment = Alignment(horizontal="center", vertical="center")
            rank_cell.font = Font(bold=True, size=10)
            
            if idx <= 3:
                rank_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            user_cell = ws.cell(row=row, column=2, value=item.get('User Name', ''))
            user_cell.font = Font(bold=True, size=10)
            
            ws.cell(row=row, column=3, value=item.get('Total Tasks', 0)).number_format = '#,##0'
            ws.cell(row=row, column=4, value=item.get('Eligible', 0)).number_format = '#,##0'
            ws.cell(row=row, column=5, value=item.get('Not Eligible', 0)).number_format = '#,##0'
            
            success_formula = f"=IF(C{row}>0,D{row}/C{row},0)"
            success_cell = ws.cell(row=row, column=6)
            success_cell.value = success_formula
            success_cell.number_format = '0.0%'
            success_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=row, column=7, value=item.get('Invited', 0)).number_format = '#,##0'
            ws.cell(row=row, column=8, value=item.get('Changed Mind', 0)).number_format = '#,##0'
            ws.cell(row=row, column=9, value=item.get('No Response', 0)).number_format = '#,##0'
            
            response_formula = f"=IF(C{row}>0,(G{row}+H{row}+I{row})/C{row},0)"
            response_rate_cell = ws.cell(row=row, column=10)
            response_rate_cell.value = response_formula
            response_rate_cell.number_format = '0.0%'
            response_rate_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            conversion_formula = f"=IF(G{row}>0,D{row}/G{row},0)"
            conversion_rate_cell = ws.cell(row=row, column=11)
            conversion_rate_cell.value = conversion_formula
            conversion_rate_cell.number_format = '0.0%'
            conversion_rate_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            perf_level_formula = f'=IF(F{row}>=0.7,"Excellent",IF(F{row}>=0.5,"Good",IF(F{row}>=0.3,"Fair","Needs Improvement")))'
            level_cell = ws.cell(row=row, column=12)
            level_cell.value = perf_level_formula
            level_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.cell(row=row, column=13, value=item.get('Projects', '')).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=14, value=item.get('Cities', '')).alignment = Alignment(wrap_text=True)
            
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB")
                )
        
        column_widths = [7, 20, 12, 10, 12, 14, 10, 13, 12, 14, 14, 18, 25, 25]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        for i in range(4, 4 + len(sorted_performance) + 1):
            ws.row_dimensions[i].height = 24
        
        ws.freeze_panes = "C5"
        
        chart_row = len(sorted_performance) + 7
        self.create_performance_charts(ws, sorted_performance, chart_row)
    
    def create_performance_charts(self, ws, performance_data, start_row):
        ws.cell(row=start_row, column=1, value="PERFORMANCE VISUALIZATION").font = Font(bold=True, size=12, color=self.primary_color)
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 11
        chart1.title = "Top 15 Users by Success Rate"
        chart1.y_axis.title = 'Success Rate (%)'
        chart1.x_axis.title = 'User'
        
        data_range = Reference(ws, min_col=6, min_row=4, max_row=min(19, len(performance_data) + 4))
        categories = Reference(ws, min_col=2, min_row=5, max_row=min(19, len(performance_data) + 4))
        
        chart1.add_data(data_range, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.height = 14
        chart1.width = 22
        
        ws.add_chart(chart1, f"A{start_row + 2}")
        
        level_data_row = start_row + 22
        ws.cell(row=level_data_row, column=1, value="Performance Level Distribution").font = Font(bold=True, size=11)
        ws.cell(row=level_data_row + 1, column=1, value="Level").font = Font(bold=True)
        ws.cell(row=level_data_row + 1, column=2, value="Count").font = Font(bold=True)
        
        levels = ["Excellent", "Good", "Fair", "Needs Improvement"]
        for i, level in enumerate(levels, 1):
            ws.cell(row=level_data_row + 1 + i, column=1, value=level)
            count_formula = f'=COUNTIF(L5:L{4 + len(performance_data)},A{level_data_row + 1 + i})'
            ws.cell(row=level_data_row + 1 + i, column=2, value=count_formula)
        
        pie_chart = PieChart()
        pie_chart.title = "Distribution by Performance Level"
        
        pie_data = Reference(ws, min_col=2, min_row=level_data_row + 1, max_row=level_data_row + 1 + len(levels))
        pie_categories = Reference(ws, min_col=1, min_row=level_data_row + 2, max_row=level_data_row + 1 + len(levels))
        
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        pie_chart.height = 13
        pie_chart.width = 18
        
        ws.add_chart(pie_chart, f"L{start_row + 2}")
    
    def create_insights_sheet(self, wb, data):
        ws = wb.create_sheet("Strategic Insights")
        insights_data = data.get('insightsData', [])
        
        if not insights_data:
            return
        
        title = ws.cell(row=1, column=1, value="STRATEGIC INSIGHTS & PRIORITY AREAS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:H1")
        
        headers = ["Category", "User", "Total Tasks", "Success Rate %", "Performance Level", "Rank", "Key Issues", "Priority"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        top_performers = [item for item in insights_data if item.get('Category') == 'Top Performer']
        priority_areas = [item for item in insights_data if item.get('Category') == 'Priority Area']
        volume_leaders = [item for item in insights_data if item.get('Category') == 'Volume Leader']
        
        current_row = 4
        
        categories_config = [
            ("TOP PERFORMERS - EXCELLENCE IN EXECUTION", top_performers, "D1FAE5", self.success_color),
            ("PRIORITY AREAS - IMMEDIATE ATTENTION REQUIRED", priority_areas, "FEE2E2", self.danger_color),
            ("VOLUME LEADERS - HIGH PRODUCTIVITY", volume_leaders, "DBEAFE", self.secondary_color)
        ]
        
        for category_name, category_data, bg_color, text_color in categories_config:
            if not category_data:
                continue
            
            category_cell = ws.cell(row=current_row, column=1, value=category_name)
            category_cell.font = Font(bold=True, size=12, color=text_color)
            category_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            category_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws.row_dimensions[current_row].height = 28
            current_row += 1
            
            for item in category_data:
                ws.cell(row=current_row, column=1, value=item.get('Category', ''))
                ws.cell(row=current_row, column=2, value=item.get('User', ''))
                
                tasks_cell = ws.cell(row=current_row, column=3, value=item.get('Total Tasks', 0))
                tasks_cell.number_format = '#,##0'
                tasks_cell.alignment = Alignment(horizontal="right", vertical="center")
                
                orig_index = item.get('OriginalIndex', 0)
                perf_sheet_row = orig_index + 5
                
                success_formula = f"='User Performance Analysis'!F{perf_sheet_row}"
                success_cell = ws.cell(row=current_row, column=4)
                success_cell.value = success_formula
                success_cell.number_format = '0.0%'
                success_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                perf_level_formula = f"='User Performance Analysis'!L{perf_sheet_row}"
                perf_cell = ws.cell(row=current_row, column=5)
                perf_cell.value = perf_level_formula
                perf_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                rank_cell = ws.cell(row=current_row, column=6, value=item.get('Rank', ''))
                rank_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                issues_cell = ws.cell(row=current_row, column=7, value=item.get('Issues', '-'))
                issues_cell.alignment = Alignment(wrap_text=True, vertical="center")
                issues_cell.font = Font(size=9)
                
                priority = "High" if item.get('Category') == 'Priority Area' else "Medium" if item.get('Category') == 'Volume Leader' else "Normal"
                priority_cell = ws.cell(row=current_row, column=8, value=priority)
                priority_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if priority == "High":
                    priority_cell.fill = PatternFill(start_color=self.danger_color, end_color=self.danger_color, fill_type="solid")
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                elif priority == "Medium":
                    priority_cell.fill = PatternFill(start_color=self.warning_color, end_color=self.warning_color, fill_type="solid")
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = Border(
                        left=Side(style="thin", color="D1D5DB"),
                        right=Side(style="thin", color="D1D5DB"),
                        top=Side(style="thin", color="D1D5DB"),
                        bottom=Side(style="thin", color="D1D5DB")
                    )
                
                current_row += 1
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 12
        
        ws.freeze_panes = "A4"
        
        if top_performers:
            chart_row = current_row + 2
            self.create_insights_charts(ws, top_performers, chart_row)
    
    def create_insights_charts(self, ws, top_performers, start_row):
        ws.cell(row=start_row, column=1, value="PERFORMANCE COMPARISON").font = Font(bold=True, size=12, color=self.primary_color)
        
        chart_data_row = start_row + 2
        ws.cell(row=chart_data_row, column=1, value="User").font = Font(bold=True)
        ws.cell(row=chart_data_row, column=2, value="Success Rate (%)").font = Font(bold=True)
        
        for i, item in enumerate(top_performers[:10], 1):
            ws.cell(row=chart_data_row + i, column=1, value=item.get('User', ''))
            
            orig_index = item.get('OriginalIndex', 0)
            perf_sheet_row = orig_index + 5
            success_formula = f"='User Performance Analysis'!F{perf_sheet_row}"
            ws.cell(row=chart_data_row + i, column=2, value=success_formula)
        
        chart = BarChart()
        chart.type = "bar"
        chart.style = 13
        chart.title = "Top 10 Performers by Success Rate"
        chart.x_axis.title = 'Success Rate (%)'
        chart.y_axis.title = 'User'
        
        chart_data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + min(len(top_performers), 10))
        chart_categories = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + min(len(top_performers), 10))
        
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_categories)
        chart.height = 13
        chart.width = 20
        
        ws.add_chart(chart, f"A{start_row + 4}")
    
    def create_recommendations_sheet(self, wb, data):
        ws = wb.create_sheet("Management Recommendations")
        
        title = ws.cell(row=1, column=1, value="MANAGEMENT RECOMMENDATIONS & ACTION ITEMS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        ws.cell(row=2, column=1, value="Data-Driven Insights for Leadership Decision Making").font = Font(size=11, color="6B7280", italic=True)
        
        performance_data = data.get('performanceData', [])
        summary_data = data.get('summaryData', [])
        insights_data = data.get('insightsData', [])
        
        current_row = 4
        
        recommendations = []
        
        total_tasks = next((item['Value'] for item in summary_data if item['Metric'] == 'Total Tasks'), 0)
        total_eligible = next((item['Value'] for item in summary_data if item['Metric'] == 'Total Eligible'), 0)
        
        if total_tasks > 0:
            overall_success = (total_eligible / total_tasks) * 100
            if overall_success < 50:
                recommendations.append({
                    'category': 'Critical',
                    'title': 'Overall Success Rate Below Target',
                    'description': f'Current success rate requires immediate intervention.',
                    'actions': [
                        'Conduct comprehensive review of task qualification criteria',
                        'Implement mandatory training program for all users',
                        'Establish weekly performance monitoring meetings',
                        'Review and update standard operating procedures'
                    ],
                    'timeline': 'Immediate (1-2 weeks)',
                    'owner': 'Operations Manager'
                })
        
        priority_users = [item for item in insights_data if item.get('Category') == 'Priority Area']
        if len(priority_users) > 0:
            recommendations.append({
                'category': 'High Priority',
                'title': f'Performance Improvement Required for {len(priority_users)} Users',
                'description': 'Multiple users showing below-target performance metrics requiring immediate coaching.',
                'actions': [
                    'Schedule one-on-one coaching sessions with underperforming users',
                    'Assign mentors from top performer group',
                    'Implement 30-day performance improvement plans',
                    'Provide additional resources and training materials'
                ],
                'timeline': 'Short-term (2-4 weeks)',
                'owner': 'Team Lead & HR'
            })
        
        top_performers = [item for item in insights_data if item.get('Category') == 'Top Performer']
        if len(top_performers) >= 3:
            recommendations.append({
                'category': 'Strategic',
                'title': 'Leverage Top Performers for Knowledge Transfer',
                'description': f'{len(top_performers)} users demonstrating excellence in task execution.',
                'actions': [
                    'Document best practices from top performers',
                    'Establish peer mentoring program',
                    'Create case studies of successful task completions',
                    'Recognize and reward top performers publicly'
                ],
                'timeline': 'Medium-term (1-2 months)',
                'owner': 'Training & Development'
            })
        
        total_invited = next((item['Value'] for item in summary_data if item['Metric'] == 'Total Invited'), 0)
        total_changed_mind = next((item['Value'] for item in summary_data if item['Metric'] == 'Total Changed Mind'), 0)
        total_no_response = next((item['Value'] for item in summary_data if item['Metric'] == 'Total No Response'), 0)
        
        if total_tasks > 0:
            response_rate = ((total_invited + total_changed_mind + total_no_response) / total_tasks) * 100
            if response_rate < 70:
                recommendations.append({
                    'category': 'Operational',
                    'title': 'Improve Response Rate and Follow-up Process',
                    'description': 'Response rate optimization opportunity identified.',
                    'actions': [
                        'Implement automated follow-up reminder system',
                        'Review and optimize contact timing strategies',
                        'Develop multi-channel communication approach',
                        'Create response tracking dashboard'
                    ],
                    'timeline': 'Medium-term (4-6 weeks)',
                    'owner': 'Operations Team'
                })
        
        recommendations.append({
            'category': 'Strategic',
            'title': 'Data-Driven Decision Making Enhancement',
            'description': 'Establish robust analytics framework for continuous improvement.',
            'actions': [
                'Implement real-time performance monitoring dashboard',
                'Schedule monthly analytics review meetings',
                'Develop predictive models for success rate optimization',
                'Create automated reporting system for management'
            ],
            'timeline': 'Long-term (2-3 months)',
            'owner': 'Data Analytics Team'
        })
        
        for rec in recommendations:
            cat_cell = ws.cell(row=current_row, column=1, value=f"[{rec['category'].upper()}]")
            cat_cell.font = Font(bold=True, size=11, color="FFFFFF")
            
            if rec['category'] == 'Critical':
                cat_cell.fill = PatternFill(start_color=self.danger_color, end_color=self.danger_color, fill_type="solid")
            elif rec['category'] == 'High Priority':
                cat_cell.fill = PatternFill(start_color=self.warning_color, end_color=self.warning_color, fill_type="solid")
            else:
                cat_cell.fill = PatternFill(start_color=self.secondary_color, end_color=self.secondary_color, fill_type="solid")
            
            cat_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            
            title_cell = ws.cell(row=current_row, column=1, value=rec['title'])
            title_cell.font = Font(bold=True, size=12)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            current_row += 1
            
            desc_cell = ws.cell(row=current_row, column=1, value=rec['description'])
            desc_cell.font = Font(size=10, color="374151", italic=True)
            desc_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            action_header = ws.cell(row=current_row, column=1, value="Action Items:")
            action_header.font = Font(bold=True, size=10)
            current_row += 1
            
            for action in rec['actions']:
                action_cell = ws.cell(row=current_row, column=1, value=f"  • {action}")
                action_cell.font = Font(size=10)
                action_cell.alignment = Alignment(wrap_text=True, indent=1)
                ws.merge_cells(f"A{current_row}:F{current_row}")
                ws.row_dimensions[current_row].height = 25
                current_row += 1
            
            timeline_cell = ws.cell(row=current_row, column=1, value=f"Timeline: {rec['timeline']}")
            timeline_cell.font = Font(size=10, bold=True, color=self.purple_color)
            ws.merge_cells(f"A{current_row}:C{current_row}")
            
            owner_cell = ws.cell(row=current_row, column=4, value=f"Owner: {rec['owner']}")
            owner_cell.font = Font(size=10, bold=True, color=self.purple_color)
            ws.merge_cells(f"D{current_row}:F{current_row}")
            current_row += 2
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

def main():
    try:
        if len(sys.argv) != 3:
            raise ValueError("Usage: python taskAnalyticsChartGenerator.py <input_json_path> <output_excel_path>")
        
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        generator = TaskAnalyticsChartGenerator()
        result_path = generator.create_workbook_with_charts(data, output_path)
        
        print(json.dumps({
            "success": True,
            "output_path": result_path,
            "message": "Task performance analytics dashboard created successfully"
        }))
    
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()