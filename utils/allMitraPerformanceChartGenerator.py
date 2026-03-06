import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, PieChart, RadarChart, AreaChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class AllMitraPerformanceChartGenerator:
    def __init__(self):
        self.primary_color = "1E3A8A"
        self.secondary_color = "3B82F6"
        self.success_color = "10B981"
        self.warning_color = "F59E0B"
        self.danger_color = "EF4444"
        self.light_bg = "F3F4F6"
        self.header_bg = "1E40AF"
        
    def create_workbook_with_charts(self, data, output_path):
        wb = openpyxl.Workbook()
        
        period_type = data.get('periodType', 'monthly')
        
        self.create_constants_sheet(wb)
        self.create_overview_sheet(wb, data, period_type)
        self.create_performance_metrics_sheet(wb, data, period_type)
        self.create_cost_analysis_sheet(wb, data, period_type)
        self.create_top_performers_sheet(wb, data, period_type)
        self.create_city_distribution_sheet(wb, data, period_type)
        self.create_performance_trends_sheet(wb, data, period_type)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.active = wb['Overview']
        wb.save(output_path)
        return output_path
    
    def create_constants_sheet(self, wb):
        ws = wb.create_sheet("Constants", 0)
        ws.sheet_state = 'hidden'
        
        title = ws.cell(row=1, column=1, value="SYSTEM CONSTANTS")
        title.font = Font(bold=True, size=14, color=self.primary_color)
        
        ws.cell(row=3, column=1, value="Constant Name").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Value").font = Font(bold=True)
        ws.cell(row=3, column=3, value="Description").font = Font(bold=True)
        
        constants = [
            ("DELIVERY_RATE_TARGET", 95, "Target delivery success rate (95%)"),
            ("ONTIME_RATE_TARGET", 90, "Target on-time delivery rate (90%)"),
            ("ACTIVITY_BASELINE", 100, "Baseline for activity level calculation"),
            ("WEIGHT_DELIVERY_RATE", 0.40, "Weight for delivery rate in score (40%)"),
            ("WEIGHT_ONTIME_RATE", 0.30, "Weight for on-time rate in score (30%)"),
            ("WEIGHT_ACTIVITY", 0.20, "Weight for activity level in score (20%)"),
            ("WEIGHT_CONSISTENCY", 0.10, "Weight for consistency in score (10%)")
        ]
        
        for idx, (name, value, desc) in enumerate(constants, 4):
            ws.cell(row=idx, column=1, value=name).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
            ws.cell(row=idx, column=3, value=desc).font = Font(italic=True, size=9)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
    
    def create_overview_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Overview", 1)
        
        title = ws.cell(row=2, column=2, value=f"ALL MITRA PERFORMANCE OVERVIEW - {period_type.upper()}")
        title.font = Font(bold=True, size=20, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:H2")
        ws.row_dimensions[2].height = 35
        
        subtitle = ws.cell(row=3, column=2, value="Comprehensive Performance Analytics for All Mitra Partners")
        subtitle.font = Font(size=12, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("B3:H3")
        
        ws.cell(row=5, column=2, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=14, color=self.primary_color)
        
        mitras = data.get('mitras', [])
        total_mitras = len(mitras)
        total_deliveries = sum(m.get('totalDeliveries', 0) for m in mitras)
        total_cost = sum(m.get('totalCost', 0) for m in mitras)
        avg_ontime_rate = sum(m.get('onTimeRate', 0) for m in mitras) / total_mitras if total_mitras > 0 else 0
        
        kpi_data = [
            ("Total Mitra Partners", total_mitras, "partners"),
            ("Total Deliveries", total_deliveries, "deliveries"),
            ("Total Cost", total_cost, "IDR"),
            ("Average On-Time Rate", avg_ontime_rate, "%")
        ]
        
        row = 7
        for label, value, unit in kpi_data:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=12)
            ws.cell(row=row, column=3, value=value).font = Font(size=14, bold=True, color=self.secondary_color)
            ws.cell(row=row, column=4, value=unit).font = Font(size=10, italic=True)
            row += 1
        
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
    
    def create_performance_metrics_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Performance Metrics", 2)
        
        title = ws.cell(row=1, column=1, value=f"PERFORMANCE METRICS - {period_type.upper()}")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        headers = ["Mitra Name", "Total Deliveries", "On-Time Rate", "Avg Cost", "Avg Distance", "Cost per Km"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        mitras = data.get('mitras', [])
        for idx, mitra in enumerate(mitras, 4):
            ws.cell(row=idx, column=1, value=mitra.get('name', ''))
            ws.cell(row=idx, column=2, value=mitra.get('totalDeliveries', 0))
            ws.cell(row=idx, column=3, value=mitra.get('onTimeRate', 0)).number_format = '0.00%'
            ws.cell(row=idx, column=4, value=mitra.get('avgCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=5, value=mitra.get('avgDistance', 0)).number_format = '0.00'
            ws.cell(row=idx, column=6, value=mitra.get('costPerKm', 0)).number_format = '#,##0'
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_cost_analysis_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Cost Analysis", 3)
        
        title = ws.cell(row=1, column=1, value=f"COST ANALYSIS - {period_type.upper()}")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        headers = ["Mitra Name", "Total Cost", "Avg Cost per Delivery", "Total Distance", "Avg Distance", "Cost per Km"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        mitras = data.get('mitras', [])
        for idx, mitra in enumerate(mitras, 4):
            ws.cell(row=idx, column=1, value=mitra.get('name', ''))
            ws.cell(row=idx, column=2, value=mitra.get('totalCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=3, value=mitra.get('avgCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=4, value=mitra.get('totalDistance', 0)).number_format = '0.00'
            ws.cell(row=idx, column=5, value=mitra.get('avgDistance', 0)).number_format = '0.00'
            ws.cell(row=idx, column=6, value=mitra.get('costPerKm', 0)).number_format = '#,##0'
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_top_performers_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Top Performers", 4)
        
        title = ws.cell(row=1, column=1, value=f"TOP PERFORMERS - {period_type.upper()}")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        headers = ["Rank", "Mitra Name", "Total Deliveries", "On-Time Rate", "Total Cost", "Avg Cost"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        mitras = data.get('mitras', [])
        top_performers = sorted(mitras, key=lambda x: x.get('totalDeliveries', 0), reverse=True)[:20]
        
        for idx, mitra in enumerate(top_performers, 4):
            ws.cell(row=idx, column=1, value=idx - 3)
            ws.cell(row=idx, column=2, value=mitra.get('name', ''))
            ws.cell(row=idx, column=3, value=mitra.get('totalDeliveries', 0))
            ws.cell(row=idx, column=4, value=mitra.get('onTimeRate', 0)).number_format = '0.00%'
            ws.cell(row=idx, column=5, value=mitra.get('totalCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=6, value=mitra.get('avgCost', 0)).number_format = '#,##0'
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_city_distribution_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("City Distribution", 5)
        
        title = ws.cell(row=1, column=1, value=f"CITY DISTRIBUTION - {period_type.upper()}")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:D1")
        
        headers = ["City", "Mitra Count", "Total Deliveries", "Avg On-Time Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        mitras = data.get('mitras', [])
        city_stats = {}
        
        for mitra in mitras:
            hubs = mitra.get('hubs', [])
            if hubs and isinstance(hubs, list):
                for hub in hubs:
                    if hub and hub.strip():
                        if hub not in city_stats:
                            city_stats[hub] = {
                                'count': 0,
                                'deliveries': 0,
                                'ontime_rates': []
                            }
                        city_stats[hub]['count'] += 1
                        city_stats[hub]['deliveries'] += mitra.get('totalDeliveries', 0)
                        city_stats[hub]['ontime_rates'].append(mitra.get('onTimeRate', 0))
        
        sorted_cities = sorted(city_stats.items(), key=lambda x: x[1]['deliveries'], reverse=True)[:20]
        
        for idx, (city, stats) in enumerate(sorted_cities, 4):
            avg_ontime = sum(stats['ontime_rates']) / len(stats['ontime_rates']) if stats['ontime_rates'] else 0
            ws.cell(row=idx, column=1, value=city)
            ws.cell(row=idx, column=2, value=stats['count'])
            ws.cell(row=idx, column=3, value=stats['deliveries'])
            ws.cell(row=idx, column=4, value=avg_ontime).number_format = '0.00%'
        
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_performance_trends_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Performance Trends", 6)
        
        title = ws.cell(row=1, column=1, value=f"PERFORMANCE TRENDS - {period_type.upper()}")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        headers = ["Mitra Name", "Total Deliveries", "On-Time Rate", "Total Cost", "Avg Cost", "Cost per Km"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        mitras = data.get('mitras', [])
        sorted_by_deliveries = sorted(mitras, key=lambda x: x.get('totalDeliveries', 0), reverse=True)
        
        for idx, mitra in enumerate(sorted_by_deliveries[:50], 4):
            ws.cell(row=idx, column=1, value=mitra.get('name', ''))
            ws.cell(row=idx, column=2, value=mitra.get('totalDeliveries', 0))
            ws.cell(row=idx, column=3, value=mitra.get('onTimeRate', 0)).number_format = '0.00%'
            ws.cell(row=idx, column=4, value=mitra.get('totalCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=5, value=mitra.get('avgCost', 0)).number_format = '#,##0'
            ws.cell(row=idx, column=6, value=mitra.get('costPerKm', 0)).number_format = '#,##0'
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

def main():
    try:
        if len(sys.argv) != 3:
            raise ValueError("Usage: python allMitraPerformanceChartGenerator.py <input_json> <output_excel>")
        
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        mitras = data.get('mitras', [])
        period_type = data.get('periodType', 'monthly')
        
        if len(mitras) == 0:
            print(json.dumps({
                "success": False,
                "error": "No mitra data available. Cannot generate report without mitra data.",
                "details": "Please ensure mitra data is available before generating the report."
            }))
            sys.exit(1)
        
        generator = AllMitraPerformanceChartGenerator()
        result_path = generator.create_workbook_with_charts(data, output_path)
        
        message = f"All mitra performance chart with {period_type} data created successfully"
        
        print(json.dumps({
            "success": True,
            "output_path": result_path,
            "message": message,
            "data_summary": {
                "total_mitras": len(mitras),
                "period_type": period_type
            }
        }))
    
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()