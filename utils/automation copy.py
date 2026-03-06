from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import time
import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import warnings
warnings.filterwarnings('ignore')

ADMINPANEL_STATUS = os.getenv("ADMINPANEL_STATUS", "false").lower() == "true"
ADMINMANAGE_STATUS = os.getenv("ADMINMANAGE_STATUS", "false").lower() == "true"


class GoogleSheetsDownloader:
    TEMPLATE_URL = "https://drive.google.com/uc?export=download&id=1c5W-93qD-TK7zMvMl994yT6Au9jdxTYf"

    REQUIRED_COLUMNS = [
        'merchant_order_id*', 'weight*', 'width', 'height', 'length',
        'payment_type*', 'cod_amount', 'sender_name*', 'sender_phone*',
        'pickup_instructions', 'consignee_name*', 'consignee_phone*',
        'destination_district', 'destination_city*', 'destination_province',
        'destination_postalcode*', 'destination_address*', 'dropoff_lat',
        'dropoff_long', 'dropoff_instructions', 'item_value*', 'product_details*'
    ]

    PHONE_COLUMNS = [8, 11]

    def __init__(self, sheet_url, worksheet_name="OPERATIONS"):
        self.sheet_url = sheet_url
        self.worksheet_name = worksheet_name
        self.sheet_id = self._extract_sheet_id(sheet_url)

    def _extract_sheet_id(self, url):
        if "/d/" in url:
            return url.split("/d/")[1].split("/")[0]
        return url

    def _get_gid_from_url(self, url):
        if "#gid=" in url:
            return url.split("#gid=")[1].split("&")[0]
        return "0"

    def _clean_phone_number(self, value):
        if pd.isna(value) or value == '':
            return ''
        value_str = str(value)
        if 'E+' in value_str or 'e+' in value_str:
            try:
                float_val = float(value_str)
                return f"{int(float_val)}"
            except:
                pass
        return value_str.replace('.0', '').replace(',', '').replace(' ', '')

    def download_template(self, template_path):
        try:
            response = requests.get(self.TEMPLATE_URL, timeout=30)
            response.raise_for_status()
            with open(template_path, 'wb') as f:
                f.write(response.content)
            return True
        except Exception:
            return False

    def download_as_excel(self, output_path):
        gid = self._get_gid_from_url(self.sheet_url)
        export_url = f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/export?format=csv&gid={gid}"

        response = requests.get(export_url, timeout=30)
        response.raise_for_status()

        import io
        csv_content = response.content.decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_content), dtype=str, keep_default_na=False)

        if len(df.columns) == len(self.REQUIRED_COLUMNS):
            df.columns = self.REQUIRED_COLUMNS

        for col_idx in self.PHONE_COLUMNS:
            if col_idx < len(df.columns):
                df.iloc[:, col_idx] = df.iloc[:, col_idx].apply(self._clean_phone_number)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        template_path = os.path.join(os.path.dirname(output_path), "template.xlsx")

        if not os.path.exists(template_path):
            if not self.download_template(template_path):
                df.to_excel(output_path, index=False, sheet_name='Sheet1', engine='openpyxl')
                wb = load_workbook(output_path)
                ws = wb.active
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=10, name='Calibri')
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                wb.save(output_path)
                wb.close()
                return output_path

        wb_template = load_workbook(template_path)
        ws = wb_template.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                cell.value = '' if (value == '' or pd.isna(value)) else str(value)

        wb_template.save(output_path)
        wb_template.close()

        return output_path


class BlitzAutomation:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.login_url = "https://adminpanel.rideblitz.id/login/"
        self.base_form_url = "https://adminpanel.rideblitz.id/api/bulkorderactivity/add/"
        self.admin_manage_base_url = "https://admin-manage.rideblitz.id"

    def _build_form_url(self, business, city, service_type):
        return f"{self.base_form_url}?business={business}&city={city}&service_type={service_type}"

    def setup_driver(self, headless=True):
        options = webdriver.ChromeOptions()

        if headless:
            options.add_argument('--headless=new')

        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        print(f"[BROWSER] Mode: {'VISIBLE' if not headless else 'HEADLESS'}")

        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 30)

    def login(self, username, password):
        self.driver.get(self.login_url)
        username_field = self.wait.until(EC.presence_of_element_located((By.ID, "id_username")))
        password_field = self.driver.find_element(By.ID, "id_password")
        username_field.clear()
        username_field.send_keys(username)
        password_field.clear()
        password_field.send_keys(password)
        self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
        time.sleep(2)

        if ADMINPANEL_STATUS:
            print(f"[DEBUG] URL setelah login: {self.driver.current_url}")
            print(f"[DEBUG] Title: {self.driver.title}")

    def fill_bulk_order_form(self, file_path, business_hub_value=None, business=12, city=9, service_type=2):
        form_url = self._build_form_url(business, city, service_type)
        self.driver.get(form_url)

        if ADMINPANEL_STATUS:
            print(f"[DEBUG] Form URL: {form_url}")

        self.wait.until(EC.presence_of_element_located((By.ID, "bulkorderactivity_form")))

        business_hub_select = Select(self.driver.find_element(By.ID, "id_business_hub"))
        business_hub_select.select_by_value(str(business_hub_value) if business_hub_value else "59")

        if ADMINPANEL_STATUS:
            selected = business_hub_select.first_selected_option
            print(f"[DEBUG] Business hub: {selected.text} (value={selected.get_attribute('value')})")

        midmile_checkbox = self.driver.find_element(By.ID, "id_midmile_required")
        if midmile_checkbox.is_selected():
            midmile_checkbox.click()

        if not file_path or not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        file_input = self.driver.find_element(By.ID, "id_file")
        file_input.send_keys(os.path.abspath(file_path))
        time.sleep(3)

        if ADMINPANEL_STATUS:
            print(f"[DEBUG] File di-upload: {file_path}")

        self.wait.until(EC.element_to_be_clickable((By.ID, "save_btn")))

    def _screenshot(self, label, context="adminpanel"):
        debug_active = ADMINPANEL_STATUS if context == "adminpanel" else ADMINMANAGE_STATUS
        if debug_active:
            try:
                path = f"/tmp/blitz_{context}_{label}_{int(time.time())}.png"
                self.driver.save_screenshot(path)
                print(f"[DEBUG][{context.upper()}] Screenshot: {path}")
            except Exception as e:
                print(f"[DEBUG][{context.upper()}] Screenshot gagal: {e}")

    def _log_all_buttons(self, context="adminpanel"):
        debug_active = ADMINPANEL_STATUS if context == "adminpanel" else ADMINMANAGE_STATUS
        if not debug_active:
            return
        try:
            print(f"[DEBUG][{context.upper()}] URL: {self.driver.current_url}")
            buttons = self.driver.find_elements(By.TAG_NAME, "button")
            print(f"[DEBUG][{context.upper()}] Total button: {len(buttons)}")
            for btn in buttons:
                try:
                    print(f"[DEBUG][{context.upper()}]   button | text='{btn.text.strip()}' id='{btn.get_attribute('id')}' type='{btn.get_attribute('type')}' visible={btn.is_displayed()}")
                except Exception:
                    pass
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            for inp in inputs:
                try:
                    if inp.get_attribute("type") in ["submit", "button"]:
                        print(f"[DEBUG][{context.upper()}]   input  | type='{inp.get_attribute('type')}' value='{inp.get_attribute('value')}' id='{inp.get_attribute('id')}' visible={inp.is_displayed()}")
                except Exception:
                    pass
            links = self.driver.find_elements(By.TAG_NAME, "a")
            for link in links:
                try:
                    text = link.text.strip()
                    if text and link.is_displayed():
                        print(f"[DEBUG][{context.upper()}]   link   | text='{text}' href='{link.get_attribute('href')}'")
                except Exception:
                    pass
        except Exception as e:
            print(f"[DEBUG][{context.upper()}] Log buttons error: {e}")

    def submit_form(self):
        save_button = self.wait.until(EC.presence_of_element_located((By.ID, "save_btn")))
        self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", save_button)
        time.sleep(1)
        self.wait.until(EC.element_to_be_clickable((By.ID, "save_btn")))

        self._screenshot("before_save", "adminpanel")
        self.driver.execute_script("arguments[0].click();", save_button)

        time.sleep(10)

        self._screenshot("after_save", "adminpanel")

        confirm_selectors = [
            (By.XPATH, "//button[contains(text(), 'Confirm and Submit')]"),
            (By.XPATH, "//button[contains(text(), 'Confirm')]"),
            (By.XPATH, "//button[contains(text(), 'Submit')]"),
            (By.XPATH, "//input[@type='submit']"),
            (By.CSS_SELECTOR, "button.confirm-btn"),
            (By.CSS_SELECTOR, "button[type='submit']"),
            (By.XPATH, "//button[contains(@class, 'confirm')]"),
            (By.XPATH, "//a[contains(text(), 'Confirm')]"),
        ]

        confirm_button = None
        found_selector = None

        for selector_type, selector_value in confirm_selectors:
            try:
                elements = self.driver.find_elements(selector_type, selector_value)
                for el in elements:
                    if el.is_displayed() and el.is_enabled():
                        confirm_button = el
                        found_selector = f"{selector_type}={selector_value}"
                        break
                if confirm_button:
                    break
            except Exception:
                continue

        if ADMINPANEL_STATUS:
            print(f"[DEBUG][ADMINPANEL] Confirm button: {found_selector if confirm_button else 'TIDAK DITEMUKAN'}")
            if not confirm_button:
                self._log_all_buttons("adminpanel")
                self._screenshot("no_confirm_button", "adminpanel")

        if not confirm_button:
            try:
                current_url = self.driver.current_url
                if ADMINPANEL_STATUS:
                    print(f"[DEBUG][ADMINPANEL] Cek redirect URL: {current_url}")
                if "add" not in current_url:
                    if ADMINPANEL_STATUS:
                        print("[DEBUG][ADMINPANEL] Halaman sudah redirect — submission dianggap selesai")
                    return
            except Exception:
                pass

            raise RuntimeError(
                "Tombol konfirmasi tidak ditemukan. Set ADMINPANEL_STATUS=true di .env untuk melihat browser dan mendiagnosis masalah."
            )

        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_button)
            time.sleep(2)
            self.wait.until(EC.element_to_be_clickable(confirm_button))

            self._screenshot("before_confirm", "adminpanel")

            try:
                confirm_button.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", confirm_button)

            time.sleep(5)

            self._screenshot("after_confirm", "adminpanel")

            if ADMINPANEL_STATUS:
                print(f"[DEBUG][ADMINPANEL] URL setelah confirm: {self.driver.current_url}")

        except Exception as e:
            self._screenshot("confirm_error", "adminpanel")
            raise RuntimeError(f"Failed to submit confirmation: {e}")

    def navigate_admin_manage(self, path=""):
        url = f"{self.admin_manage_base_url}/{path}".rstrip("/")
        self.driver.get(url)
        time.sleep(2)

        if ADMINMANAGE_STATUS:
            print(f"[DEBUG][ADMINMANAGE] Navigated to: {self.driver.current_url}")
            print(f"[DEBUG][ADMINMANAGE] Title: {self.driver.title}")

    def open_batch_details(self, batch_id):
        path = f"batch-list/{batch_id}/batch-details"
        self.navigate_admin_manage(path)

        if ADMINMANAGE_STATUS:
            self._screenshot(f"batch_{batch_id}_details", "adminmanage")
            self._log_all_buttons("adminmanage")

        return self.driver.current_url

    def interact_admin_manage(self, batch_id, action=None):
        self.open_batch_details(batch_id)

        if ADMINMANAGE_STATUS:
            print(f"[DEBUG][ADMINMANAGE] Current URL: {self.driver.current_url}")

        if action:
            action(self.driver, self.wait)

    def close(self):
        if self.driver:
            self.driver.quit()

    def run(self, username, password, file_path=None, business_hub=None, auto_submit=False,
            google_sheet_url=None, keep_file=True, business=12, city=9, service_type=2):
        downloaded_file = None
        headless = not (ADMINPANEL_STATUS or ADMINMANAGE_STATUS)

        try:
            if google_sheet_url:
                import tempfile
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                downloads_dir = os.path.join(tempfile.gettempdir(), "blitz_downloads")
                os.makedirs(downloads_dir, exist_ok=True)
                output_file = os.path.join(downloads_dir, f"orders_{timestamp}.xlsx")
                downloader = GoogleSheetsDownloader(google_sheet_url)
                downloaded_file = downloader.download_as_excel(output_file)
                file_path = downloaded_file
            elif not file_path:
                raise ValueError("No file source configured")

            self.setup_driver(headless=headless)
            self.login(username, password)
            time.sleep(1)
            self.fill_bulk_order_form(file_path, business_hub, business=business, city=city, service_type=service_type)
            self.submit_form()

        except Exception:
            raise
        finally:
            self.close()
            if not keep_file and downloaded_file and os.path.exists(downloaded_file):
                try:
                    os.remove(downloaded_file)
                except Exception:
                    pass


if __name__ == "__main__":
    DEFAULT_GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1XFqolInJvkgT7obipcSAcZYplwncHM64Lz4MdSik4j8/edit?gid=0#gid=0"

    USERNAME = os.getenv("BLITZ_USERNAME")
    PASSWORD = os.getenv("BLITZ_PASSWORD")

    if not USERNAME or not PASSWORD:
        exit(1)

    FILE_PATH = os.getenv("BLITZ_FILE_PATH", "")
    BUSINESS_HUB = os.getenv("BLITZ_BUSINESS_HUB", "59")
    BUSINESS = int(os.getenv("BLITZ_BUSINESS", "12"))
    CITY = int(os.getenv("BLITZ_CITY", "9"))
    SERVICE_TYPE = int(os.getenv("BLITZ_SERVICE_TYPE", "2"))
    AUTO_SUBMIT = os.getenv("BLITZ_AUTO_SUBMIT", "true").lower() == "true"
    GOOGLE_SHEET_URL = os.getenv("BLITZ_GOOGLE_SHEET_URL", DEFAULT_GOOGLE_SHEET_URL)
    KEEP_FILE = os.getenv("BLITZ_KEEP_FILE", "true").lower() == "true"

    automation = BlitzAutomation()
    automation.run(
        username=USERNAME,
        password=PASSWORD,
        file_path=FILE_PATH if FILE_PATH else None,
        business_hub=BUSINESS_HUB,
        auto_submit=AUTO_SUBMIT,
        google_sheet_url=GOOGLE_SHEET_URL if GOOGLE_SHEET_URL else None,
        keep_file=KEEP_FILE,
        business=BUSINESS,
        city=CITY,
        service_type=SERVICE_TYPE
    )