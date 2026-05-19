from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import time
import os
import uuid
import sys
import json
import fcntl
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import warnings
warnings.filterwarnings('ignore')

ADMINPANEL_STATUS = os.getenv("ADMINPANEL_STATUS", "false").lower() == "true"
ADMINMANAGE_STATUS = os.getenv("ADMINMANAGE_STATUS", "false").lower() == "true"
DRIVER_ROLE = os.getenv("BLITZ_DRIVER_ROLE", "")
CHROME_LOCK_FILE = "/tmp/blitz_chrome.lock"


class ValidationError(Exception):
    def __init__(self, errors):
        self.errors = errors
        super().__init__(json.dumps(errors))


class GoogleSheetsDownloader:
    TEMPLATE_URL = "https://drive.google.com/uc?export=download&id=1c5W-93qD-TK7zMvMl994yT6Au9jdxTYf"

    REQUIRED_COLUMNS = [
        'merchant_order_id*', 'weight*', 'width', 'height', 'length',
        'payment_type*', 'cod_amount', 'sender_name*', 'sender_phone*',
        'pickup_instructions', 'consignee_name*', 'consignee_phone*',
        'destination_district', 'destination_city*', 'destination_province',
        'destination_postalcode*', 'destination_address*', 'dropoff_lat',
        'dropoff_long', 'dropoff_instructions', 'item_value*', 'product_details*'
    ]

    PHONE_COLUMNS = [8, 11]

    def __init__(self, sheet_url, worksheet_name="OPERATIONS"):
        self.sheet_url = sheet_url
        self.worksheet_name = worksheet_name
        self.sheet_id = self._extract_sheet_id(sheet_url)

    def _extract_sheet_id(self, url):
        if "/d/" in url:
            return url.split("/d/")[1].split("/")[0]
        return url

    def _get_gid_from_url(self, url):
        if "#gid=" in url:
            return url.split("#gid=")[1].split("&")[0]
        return "0"

    def _clean_phone_number(self, value):
        if pd.isna(value) or value == '':
            return ''
        value_str = str(value)
        if 'E+' in value_str or 'e+' in value_str:
            try:
                float_val = float(value_str)
                return f"{int(float_val)}"
            except Exception:
                pass
        return value_str.replace('.0', '').replace(',', '').replace(' ', '')

    def download_template(self, template_path):
        try:
            response = requests.get(self.TEMPLATE_URL, timeout=30)
            response.raise_for_status()
            with open(template_path, 'wb') as f:
                f.write(response.content)
            return True
        except Exception:
            return False

    def download_as_excel(self, output_path):
        gid = self._get_gid_from_url(self.sheet_url)
        export_url = f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/export?format=csv&gid={gid}"

        response = requests.get(export_url, timeout=30)
        response.raise_for_status()

        import io
        csv_content = response.content.decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_content), dtype=str, keep_default_na=False)

        if len(df.columns) == len(self.REQUIRED_COLUMNS):
            df.columns = self.REQUIRED_COLUMNS

        for col_idx in self.PHONE_COLUMNS:
            if col_idx < len(df.columns):
                df.iloc[:, col_idx] = df.iloc[:, col_idx].apply(self._clean_phone_number)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        template_path = os.path.join(os.path.dirname(output_path), "template.xlsx")

        if not os.path.exists(template_path):
            if not self.download_template(template_path):
                df.to_excel(output_path, index=False, sheet_name='Sheet1', engine='openpyxl')
                wb = load_workbook(output_path)
                ws = wb.active
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=10, name='Calibri')
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                wb.save(output_path)
                wb.close()
                return output_path

        wb_template = load_workbook(template_path)
        ws = wb_template.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                cell.value = '' if (value == '' or pd.isna(value)) else str(value)

        wb_template.save(output_path)
        wb_template.close()

        return output_path


class BlitzAutomation:
    def __init__(self, driver_role=None):
        role = driver_role or DRIVER_ROLE
        self._is_business = role == "business"
        base = "https://business.rideblitz.id" if self._is_business else "https://adminpanel.rideblitz.id"
        self.login_url = f"{base}/login/"
        self.base_form_url = f"{base}/api/bulkorderactivity/add/"
        self._instance_id = str(uuid.uuid4())[:8]
        self._interrupt_file = f"/tmp/blitz_interrupt_{self._instance_id}"
        self._checkpoint_file = f"/tmp/blitz_checkpoint_{self._instance_id}"
        self._lock_fd = None

    def _build_form_url(self, business, city, service_type):
        if self._is_business:
            return self.base_form_url
        return f"{self.base_form_url}?business={business}&city={city}&service_type={service_type}"

    def _acquire_chrome_lock(self):
        self._lock_fd = open(CHROME_LOCK_FILE, 'w')
        fcntl.flock(self._lock_fd, fcntl.LOCK_EX)

    def _release_chrome_lock(self):
        if self._lock_fd:
            try:
                fcntl.flock(self._lock_fd, fcntl.LOCK_UN)
                self._lock_fd.close()
            except Exception:
                pass
            self._lock_fd = None

    def _create_driver(self, headless=True):
        options = webdriver.ChromeOptions()
        if headless:
            options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument(f'--user-data-dir=/tmp/chrome_profile_{self._instance_id}')
        options.add_argument('--remote-debugging-port=0')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 30)
        return driver, wait

    def _login(self, driver, wait, username, password):
        driver.get(self.login_url)
        username_field = wait.until(EC.presence_of_element_located((By.ID, "id_username")))
        password_field = driver.find_element(By.ID, "id_password")
        username_field.clear()
        username_field.send_keys(username)
        password_field.clear()
        password_field.send_keys(password)
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
        try:
            WebDriverWait(driver, 10).until(EC.url_changes(self.login_url))
        except Exception:
            pass

    def _set_select2_value(self, driver, field_id, value):
        driver.execute_script(
            """
            var el = document.getElementById(arguments[0]);
            var opt = document.createElement('option');
            opt.value = arguments[1];
            opt.text = arguments[1];
            opt.selected = true;
            el.appendChild(opt);
            el.dispatchEvent(new Event('change', {bubbles: true}));
            """,
            field_id,
            str(value)
        )
        time.sleep(0.3)

    def _fill_bulk_order_form(self, driver, wait, file_path, business_hub_value=None, business=12, city=9, service_type=2):
        form_url = self._build_form_url(business, city, service_type)
        driver.get(form_url)

        wait.until(EC.presence_of_element_located((By.ID, "bulkorderactivity_form")))

        if self._is_business:
            self._set_select2_value(driver, "id_business", business)
            self._set_select2_value(driver, "id_city", city)
            self._set_select2_value(driver, "id_service_type", service_type)
            self._set_select2_value(driver, "id_business_hub", str(business_hub_value) if business_hub_value else "59")
        else:
            business_hub_select = Select(driver.find_element(By.ID, "id_business_hub"))
            business_hub_select.select_by_value(str(business_hub_value) if business_hub_value else "59")
            try:
                midmile_checkbox = driver.find_element(By.ID, "id_midmile_required")
                if midmile_checkbox.is_selected():
                    midmile_checkbox.click()
            except Exception:
                pass

        if not file_path or not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        file_input = driver.find_element(By.ID, "id_file")
        file_input.send_keys(os.path.abspath(file_path))

        try:
            WebDriverWait(driver, 8).until(
                lambda d: d.find_element(By.ID, "id_file").get_attribute("value") != ""
            )
        except Exception:
            time.sleep(1)

        if not self._is_business:
            wait.until(EC.element_to_be_clickable((By.ID, "save_btn")))

    def _screenshot(self, driver, label, context="adminpanel"):
        debug_active = ADMINPANEL_STATUS if context == "adminpanel" else ADMINMANAGE_STATUS
        if debug_active:
            try:
                path = f"/tmp/blitz_{context}_{self._instance_id}_{label}_{int(time.time())}.png"
                driver.save_screenshot(path)
            except Exception:
                pass

    def _is_interrupted(self):
        return os.path.exists(self._interrupt_file)

    def _mark_checkpoint(self):
        try:
            with open(self._checkpoint_file, 'w') as f:
                f.write('save_clicked')
            print(f"[CHECKPOINT][SAVE_CLICKED][{self._instance_id}]")
        except Exception:
            pass

    def _cleanup_flags(self):
        for f in [self._interrupt_file, self._checkpoint_file]:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except Exception:
                    pass

    def _wait_for_page_processing(self, driver, timeout=15):
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass
        try:
            WebDriverWait(driver, 3).until_not(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".loading, .spinner, [data-loading]"))
            )
        except Exception:
            pass

    def _extract_adminpanel_validation_errors(self, driver):
        errors = []
        try:
            error_lists = driver.find_elements(By.CSS_SELECTOR, "ul.errorlist")
            for ul in error_lists:
                items = ul.find_elements(By.TAG_NAME, "li")
                for li in items:
                    text = li.text.strip()
                    if text:
                        errors.append(text)
        except Exception:
            pass

        if not errors:
            try:
                field_errors = driver.find_elements(By.CSS_SELECTOR, "[id$='_error'] li, .field-error li, .error li")
                for el in field_errors:
                    text = el.text.strip()
                    if text and text not in errors:
                        errors.append(text)
            except Exception:
                pass

        if not errors:
            try:
                alert_errors = driver.find_elements(By.CSS_SELECTOR, ".alert-danger li, .alert-error li, .messages .error")
                for el in alert_errors:
                    text = el.text.strip()
                    if text and text not in errors:
                        errors.append(text)
            except Exception:
                pass

        return errors

    def _parse_validation_error_to_orders(self, error_messages, uploaded_orders=None):
        parsed = []
        for msg in error_messages:
            entry = {
                "validation_message": msg,
                "merchant_order_id": None,
                "field": None,
                "raw": msg,
            }

            parts = msg.split(" - ")
            if len(parts) >= 3:
                line_part = parts[0].strip()
                order_id_part = parts[1].strip() if len(parts) > 1 else None
                field_part = parts[2].strip() if len(parts) > 2 else None
                detail_part = " - ".join(parts[3:]).strip() if len(parts) > 3 else None

                entry["merchant_order_id"] = order_id_part
                entry["field"] = field_part
                if detail_part:
                    entry["validation_message"] = f"{field_part} - {detail_part}"
                else:
                    entry["validation_message"] = field_part or msg

            elif len(parts) == 2:
                entry["merchant_order_id"] = parts[0].strip()
                entry["validation_message"] = parts[1].strip()

            parsed.append(entry)

        return parsed

    def _wait_for_confirm_button(self, driver, timeout=12):
        confirm_selectors = [
            (By.XPATH, "//button[contains(text(), 'Confirm and Submit')]"),
            (By.XPATH, "//button[contains(text(), 'Confirm')]"),
            (By.XPATH, "//button[contains(text(), 'Submit')]"),
            (By.XPATH, "//input[@type='submit']"),
            (By.CSS_SELECTOR, "button.confirm-btn"),
            (By.CSS_SELECTOR, "button[type='submit']"),
            (By.XPATH, "//button[contains(@class, 'confirm')]"),
            (By.XPATH, "//a[contains(text(), 'Confirm')]"),
        ]
        deadline = time.time() + timeout
        while time.time() < deadline:
            for selector_type, selector_value in confirm_selectors:
                try:
                    els = driver.find_elements(selector_type, selector_value)
                    for el in els:
                        if el.is_displayed() and el.is_enabled():
                            return el, f"{selector_type}={selector_value}"
                except Exception:
                    continue
            try:
                if "add" not in driver.current_url:
                    return None, "redirected"
            except Exception:
                pass
            time.sleep(0.5)
        return None, None

    def _check_page_for_validation_errors(self, driver):
        errors = self._extract_adminpanel_validation_errors(driver)
        if errors:
            parsed = self._parse_validation_error_to_orders(errors)
            raise ValidationError(parsed)

    def _submit_form_business(self, driver, wait):
        submit_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit'][name='_save']")))
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", submit_btn)
        time.sleep(0.5)
        self._screenshot(driver, "before_save", "adminpanel")
        driver.execute_script("arguments[0].click();", submit_btn)

        self._wait_for_page_processing(driver, timeout=15)
        self._screenshot(driver, "after_save", "adminpanel")

        self._check_page_for_validation_errors(driver)

        self._mark_checkpoint()

        if self._is_interrupted():
            raise InterruptedError("Upload interrupted before Confirm — will be merged and re-uploaded")

        confirm_btn, found_selector = self._wait_for_confirm_button(driver, timeout=12)

        if found_selector == "redirected":
            return

        if not confirm_btn:
            self._check_page_for_validation_errors(driver)
            raise RuntimeError("Tombol Confirm and Submit tidak ditemukan di halaman konfirmasi.")

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_btn)
        time.sleep(0.5)
        self._screenshot(driver, "before_confirm", "adminpanel")
        try:
            confirm_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", confirm_btn)

        self._wait_for_page_processing(driver, timeout=10)
        self._screenshot(driver, "after_confirm", "adminpanel")

        self._check_page_for_validation_errors(driver)

    def _submit_form(self, driver, wait):
        if self._is_business:
            self._submit_form_business(driver, wait)
            return

        save_button = wait.until(EC.presence_of_element_located((By.ID, "save_btn")))
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", save_button)
        time.sleep(0.5)
        wait.until(EC.element_to_be_clickable((By.ID, "save_btn")))

        self._screenshot(driver, "before_save", "adminpanel")
        driver.execute_script("arguments[0].click();", save_button)

        self._wait_for_page_processing(driver, timeout=15)
        self._screenshot(driver, "after_save", "adminpanel")

        self._check_page_for_validation_errors(driver)

        self._mark_checkpoint()

        if self._is_interrupted():
            raise InterruptedError("Upload interrupted before Confirm — will be merged and re-uploaded")

        confirm_button, found_selector = self._wait_for_confirm_button(driver, timeout=12)

        if found_selector == "redirected":
            return

        if not confirm_button:
            self._check_page_for_validation_errors(driver)
            raise RuntimeError("Tombol konfirmasi tidak ditemukan.")

        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_button)
            time.sleep(0.5)
            wait.until(EC.element_to_be_clickable(confirm_button))
            self._screenshot(driver, "before_confirm", "adminpanel")

            try:
                confirm_button.click()
            except Exception:
                driver.execute_script("arguments[0].click();", confirm_button)

            self._wait_for_page_processing(driver, timeout=10)
            self._screenshot(driver, "after_confirm", "adminpanel")

            self._check_page_for_validation_errors(driver)

        except (InterruptedError, ValidationError):
            raise
        except Exception as e:
            self._screenshot(driver, "confirm_error", "adminpanel")
            raise RuntimeError(f"Failed to submit confirmation: {e}")

    def run(self, username, password, file_path=None, business_hub=None, auto_submit=False,
            google_sheet_url=None, keep_file=True, business=12, city=9, service_type=2):
        downloaded_file = None
        headless = not (ADMINPANEL_STATUS or ADMINMANAGE_STATUS)
        driver = None

        try:
            if google_sheet_url:
                import tempfile
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                downloads_dir = os.path.join(tempfile.gettempdir(), "blitz_downloads")
                os.makedirs(downloads_dir, exist_ok=True)
                output_file = os.path.join(downloads_dir, f"orders_{timestamp}_{self._instance_id}.xlsx")
                downloader = GoogleSheetsDownloader(google_sheet_url)
                downloaded_file = downloader.download_as_excel(output_file)
                file_path = downloaded_file
            elif not file_path:
                raise ValueError("No file source configured")

            self._acquire_chrome_lock()
            driver, wait = self._create_driver(headless=headless)
            self._login(driver, wait, username, password)
            self._fill_bulk_order_form(driver, wait, file_path, business_hub, business=business, city=city, service_type=service_type)
            self._submit_form(driver, wait)

        except ValidationError:
            raise
        except InterruptedError:
            raise SystemExit(2)
        except Exception:
            raise
        finally:
            self._release_chrome_lock()
            self._cleanup_flags()
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass
            import shutil
            profile_dir = f"/tmp/chrome_profile_{self._instance_id}"
            if os.path.exists(profile_dir):
                try:
                    shutil.rmtree(profile_dir)
                except Exception:
                    pass
            if not keep_file and downloaded_file and os.path.exists(downloaded_file):
                try:
                    os.remove(downloaded_file)
                except Exception:
                    pass


if __name__ == "__main__":
    DEFAULT_GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1XFqolInJvkgT7obipcSAcZYplwncHM64Lz4MdSik4j8/edit?gid=0#gid=0"

    USERNAME = os.getenv("BLITZ_USERNAME")
    PASSWORD = os.getenv("BLITZ_PASSWORD")

    if not USERNAME or not PASSWORD:
        exit(1)

    FILE_PATH = os.getenv("BLITZ_FILE_PATH", "")
    BUSINESS_HUB = os.getenv("BLITZ_BUSINESS_HUB", "59")
    BUSINESS = int(os.getenv("BLITZ_BUSINESS", "12"))
    CITY = int(os.getenv("BLITZ_CITY", "9"))
    SERVICE_TYPE = int(os.getenv("BLITZ_SERVICE_TYPE", "2"))
    AUTO_SUBMIT = os.getenv("BLITZ_AUTO_SUBMIT", "true").lower() == "true"
    GOOGLE_SHEET_URL = os.getenv("BLITZ_GOOGLE_SHEET_URL", DEFAULT_GOOGLE_SHEET_URL)
    KEEP_FILE = os.getenv("BLITZ_KEEP_FILE", "true").lower() == "true"

    automation = BlitzAutomation()
    try:
        automation.run(
            username=USERNAME,
            password=PASSWORD,
            file_path=FILE_PATH if FILE_PATH else None,
            business_hub=BUSINESS_HUB,
            auto_submit=AUTO_SUBMIT,
            google_sheet_url=GOOGLE_SHEET_URL if GOOGLE_SHEET_URL else None,
            keep_file=KEEP_FILE,
            business=BUSINESS,
            city=CITY,
            service_type=SERVICE_TYPE
        )
    except ValidationError as e:
        print(json.dumps({"validation_error": True, "errors": e.errors}), file=sys.stderr)
        sys.exit(3)