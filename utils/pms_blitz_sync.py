import os
import sys
import argparse
import requests
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
from dotenv import load_dotenv

load_dotenv()


class PMSBlitzIntegration:
    def __init__(self):
        internal_url = os.getenv("INTERNAL_API_URL", "http://localhost:5000")
        self.pms_api_url = f"{internal_url.rstrip('/')}/api"

        self.blitz_username = os.getenv("BLITZ_USERNAME")
        self.blitz_password = os.getenv("BLITZ_PASSWORD")
        self.mongo_uri = os.getenv("DATABASE_URI", "mongodb://localhost:27017")
        self.mongo_db = os.getenv("MONGODB_DB", "pms")

        if not self.blitz_username or not self.blitz_password:
            raise ValueError("BLITZ_USERNAME and BLITZ_PASSWORD environment variables are required")

        self._mongo_client = None

    def _get_db(self):
        if self._mongo_client is None:
            self._mongo_client = MongoClient(self.mongo_uri)
        return self._mongo_client[self.mongo_db]

    def validate_sender(self, sender_name):
        db = self._get_db()
        entry = db["adminpanel_validations"].find_one({"sender_name": sender_name})
        return entry

    def validate_senders_for_orders(self, orders):
        unique_sender_names = list({o.get("sender_name") for o in orders if o.get("sender_name")})
        db = self._get_db()

        entries = db["adminpanel_validations"].find({"sender_name": {"$in": unique_sender_names}})
        validation_map = {e["sender_name"]: e for e in entries}

        invalid = [name for name in unique_sender_names if name not in validation_map]

        return validation_map, invalid

    def get_assigned_orders(self, project, driver_ids=None):
        print(f"\n📦 Fetching assigned orders from PMS for project: {project}")
        print(f"   API URL: {self.pms_api_url}")

        try:
            response = requests.get(
                f"{self.pms_api_url}/merchant-orders/{project}/all",
                timeout=30
            )
            response.raise_for_status()

            data = response.json()
            if not data.get("success"):
                print(f"❌ Failed to fetch orders: {data.get('message')}")
                return []

            all_orders = data.get("data", [])
            assigned_orders = [o for o in all_orders if o.get("assignment_status") == "assigned"]

            if driver_ids:
                assigned_orders = [o for o in assigned_orders if o.get("assigned_to_driver_id") in driver_ids]
                print(f"   Filtered by drivers: {', '.join(driver_ids)}")

            print(f"✅ Found {len(assigned_orders)} assigned orders out of {len(all_orders)} total")
            return assigned_orders

        except Exception as e:
            print(f"❌ Error fetching orders: {e}")
            return []

    def create_excel_from_orders(self, orders):
        print(f"\n📝 Creating Excel file from {len(orders)} orders")

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_path = temp_file.name
        temp_file.close()

        wb = Workbook()
        ws = wb.active

        headers = [
            "merchant_order_id*", "weight*", "width", "height", "length",
            "payment_type*", "cod_amount", "sender_name*", "sender_phone*",
            "pickup_instructions", "consignee_name*", "consignee_phone*",
            "destination_district", "destination_city*", "destination_province",
            "destination_postalcode*", "destination_address*", "dropoff_lat",
            "dropoff_long", "dropoff_instructions", "item_value*", "product_details*"
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1).value = order.get("merchant_order_id", "")
            ws.cell(row=row_idx, column=2).value = order.get("weight", 0)
            ws.cell(row=row_idx, column=3).value = order.get("width", 0)
            ws.cell(row=row_idx, column=4).value = order.get("height", 0)
            ws.cell(row=row_idx, column=5).value = order.get("length", 0)
            ws.cell(row=row_idx, column=6).value = order.get("payment_type", "non_cod")
            ws.cell(row=row_idx, column=7).value = order.get("cod_amount", 0)
            ws.cell(row=row_idx, column=8).value = order.get("sender_name", "")
            ws.cell(row=row_idx, column=9).value = order.get("sender_phone", "")
            ws.cell(row=row_idx, column=10).value = order.get("pickup_instructions", "")
            ws.cell(row=row_idx, column=11).value = order.get("consignee_name", "")
            ws.cell(row=row_idx, column=12).value = order.get("consignee_phone", "")
            ws.cell(row=row_idx, column=13).value = order.get("destination_district", "")
            ws.cell(row=row_idx, column=14).value = order.get("destination_city", "")
            ws.cell(row=row_idx, column=15).value = order.get("destination_province", "")
            ws.cell(row=row_idx, column=16).value = order.get("destination_postalcode", "")
            ws.cell(row=row_idx, column=17).value = order.get("destination_address", "")
            ws.cell(row=row_idx, column=18).value = order.get("dropoff_lat", 0)
            ws.cell(row=row_idx, column=19).value = order.get("dropoff_long", 0)
            ws.cell(row=row_idx, column=20).value = order.get("dropoff_instructions", "")
            ws.cell(row=row_idx, column=21).value = order.get("item_value", 0)
            ws.cell(row=row_idx, column=22).value = order.get("product_details", "")

        wb.save(temp_path)
        wb.close()

        print(f"✅ Excel file created: {temp_path}")
        print(f"   Size: {os.path.getsize(temp_path)} bytes")

        return temp_path

    def get_driver_info(self, project, driver_id):
        print(f"\n👤 Fetching driver info for driver_id: {driver_id}")

        try:
            response = requests.get(
                f"{self.pms_api_url}/delivery/{project}/all",
                timeout=30
            )
            response.raise_for_status()

            data = response.json()
            if not data.get("success"):
                return None

            drivers = data.get("data", [])
            driver = next((d for d in drivers if d.get("driver_id") == str(driver_id)), None)

            if driver:
                print(f"✅ Found driver: {driver.get('driver_name')}")
            else:
                print(f"⚠ Driver {driver_id} not found")

            return driver

        except Exception as e:
            print(f"❌ Error fetching driver info: {e}")
            return None

    def run_blitz_automation(self, excel_file, driver_id, validation_entry, driver_lat=None, driver_lon=None):
        business = validation_entry.get("business", 12)
        city = validation_entry.get("city", 9)
        service_type = validation_entry.get("service_type", 2)
        hub_id = validation_entry.get("business_hub", 59)

        # Ambil config dari .env jika tersedia
        google_sheet_url = os.getenv("BLITZ_GOOGLE_SHEET_URL")
        auto_submit = os.getenv("BLITZ_AUTO_SUBMIT", "true").lower() == "true"

        print(f"\n🚀 Running Blitz automation")
        print(f"   File: {excel_file}")
        print(f"   Driver ID: {driver_id}")
        print(f"   business={business}, city={city}, service_type={service_type}, hub_id={hub_id}")

        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            automation_path = os.path.join(script_dir, "automation.py")

            if not os.path.exists(automation_path):
                print(f"❌ automation.py not found at: {automation_path}")
                return None

            sys.path.insert(0, script_dir)
            from automation import BlitzAutomation

            automation = BlitzAutomation()
            batch_id = automation.run(
                username=self.blitz_username,
                password=self.blitz_password,
                file_path=excel_file,
                business=business,
                city=city,
                service_type=service_type,
                business_hub=hub_id,
                auto_submit=auto_submit,
                google_sheet_url=google_sheet_url,
                keep_file=False
            )

            if batch_id:
                print(f"✅ Blitz automation completed with batch_id: {batch_id}")
                return batch_id

            print("❌ Blitz automation did not return batch_id")
            return None

        except Exception as e:
            print(f"❌ Blitz automation failed: {e}")
            import traceback
            traceback.print_exc()
            return None

    def update_to_created_status(self, project, order_ids, batch_id):
        print(f"\n💾 Updating orders to 'created' status with batch_id: {batch_id}")

        for order_id in order_ids:
            try:
                response = requests.put(
                    f"{self.pms_api_url}/merchant-orders/{project}/{order_id}",
                    json={"batch_id": batch_id, "assignment_status": "created"},
                    timeout=30
                )
                if response.status_code == 200:
                    print(f"   ✅ Updated {order_id} to 'created' status")
                else:
                    print(f"   ⚠ Failed to update {order_id}")
            except Exception as e:
                print(f"   ❌ Error updating {order_id}: {e}")

    def update_order_status(self, project, order_ids, batch_id=None):
        print(f"\n🔒 Locking orders with batch_id: {batch_id}")

        for order_id in order_ids:
            try:
                update_data = {"assignment_status": "in_progress"}
                if batch_id:
                    update_data["batch_id"] = batch_id

                response = requests.put(
                    f"{self.pms_api_url}/merchant-orders/{project}/{order_id}",
                    json=update_data,
                    timeout=30
                )
                if response.status_code == 200:
                    print(f"   ✅ Locked {order_id}")
                else:
                    print(f"   ⚠ Failed to lock {order_id}")
            except Exception as e:
                print(f"   ❌ Error locking {order_id}: {e}")

    def _process_sender_group(self, project, sender_name, orders, validation_entry, driver_id, driver_info):
        coordinates = validation_entry.get("location", {}).get("coordinates", [])

        print(f"\n{'='*70}")
        print(f"Processing sender: {sender_name}")
        print(f"   Orders: {len(orders)}")
        print(f"   business={validation_entry.get('business')}, city={validation_entry.get('city')}, service_type={validation_entry.get('service_type')}, hub_id={validation_entry.get('business_hub')}")
        print(f"{'='*70}")

        excel_file = self.create_excel_from_orders(orders)

        try:
            driver_lat = coordinates[1] if len(coordinates) >= 2 else (driver_info.get("lat") if driver_info else -6.212149256431801)
            driver_lon = coordinates[0] if len(coordinates) >= 2 else (driver_info.get("lon") if driver_info else 106.91958799124394)

            batch_id = self.run_blitz_automation(
                excel_file=excel_file,
                driver_id=int(driver_id),
                validation_entry=validation_entry,
                driver_lat=driver_lat,
                driver_lon=driver_lon
            )

            if batch_id:
                print(f"\n{'='*70}")
                print(f"✅ BATCH CREATED: {batch_id}")
                print(f"   URL: https://admin-manage.rideblitz.id/batch-list/{batch_id}/batch-details")
                print(f"{'='*70}")

                order_ids = [str(o["_id"]) for o in orders]
                self.update_to_created_status(project, order_ids, batch_id)
            else:
                print(f"\n❌ Batch creation failed for sender: {sender_name}")

        finally:
            if os.path.exists(excel_file):
                try:
                    os.remove(excel_file)
                    print(f"\n🗑️  Temporary file removed: {excel_file}")
                except Exception:
                    pass

    def sync_assigned_orders(self, project, driver_ids=None):
        print(f"\n{'='*70}")
        print(f"PMS TO BLITZ SYNC - PROJECT: {project.upper()}")
        print(f"{'='*70}")

        assigned_orders = self.get_assigned_orders(project, driver_ids)

        if not assigned_orders:
            print("\n⚠ No assigned orders found to sync")
            return

        print(f"\n🔍 Validating sender_name against adminpanel_validations...")
        validation_map, invalid_senders = self.validate_senders_for_orders(assigned_orders)

        if invalid_senders:
            print(f"\n❌ Sender berikut belum memiliki akun business di AdminPanel:")
            for name in invalid_senders:
                print(f"   - {name}")
            print(f"\n⚠ Proses sync dihentikan. Daftarkan sender di adminpanel_validations terlebih dahulu.")
            return

        print(f"✅ Semua sender valid")

        grouped_by_driver = {}
        for order in assigned_orders:
            driver_id = order.get("assigned_to_driver_id")
            if driver_id:
                if driver_id not in grouped_by_driver:
                    grouped_by_driver[driver_id] = []
                grouped_by_driver[driver_id].append(order)

        print(f"\n📊 Orders grouped by {len(grouped_by_driver)} driver(s)")

        for driver_id, orders in grouped_by_driver.items():
            print(f"\n{'='*70}")
            print(f"Processing Driver: {driver_id} ({len(orders)} orders)")
            print(f"{'='*70}")

            driver_info = self.get_driver_info(project, driver_id)

            grouped_by_sender = {}
            for order in orders:
                sender_name = order.get("sender_name")
                if sender_name not in grouped_by_sender:
                    grouped_by_sender[sender_name] = []
                grouped_by_sender[sender_name].append(order)

            unique_senders = list(grouped_by_sender.keys())

            if len(unique_senders) == 1:
                sender_name = unique_senders[0]
                validation_entry = validation_map[sender_name]
                self._process_sender_group(project, sender_name, orders, validation_entry, driver_id, driver_info)
            else:
                print(f"\n📦 Multiple senders detected: {len(unique_senders)}")
                for sender_name in unique_senders:
                    print(f"   - {sender_name}: {len(grouped_by_sender[sender_name])} orders")

                for sender_name, sender_orders in grouped_by_sender.items():
                    validation_entry = validation_map[sender_name]
                    self._process_sender_group(project, sender_name, sender_orders, validation_entry, driver_id, driver_info)

        print(f"\n{'='*70}")
        print("✅ SYNC COMPLETED")
        print(f"{'='*70}")


def main():
    parser = argparse.ArgumentParser(description="PMS to Blitz Integration Sync")
    parser.add_argument("--project", type=str, default="mup", help="Project name")
    parser.add_argument("--drivers", type=str, help="Comma-separated driver IDs to sync")

    args = parser.parse_args()

    driver_ids = None
    if args.drivers:
        driver_ids = [d.strip() for d in args.drivers.split(",")]

    # Baca INTERNAL_API_URL dari env untuk display
    internal_url = os.getenv("INTERNAL_API_URL", "http://localhost:5000")

    print(f"\n{'='*70}")
    print("PMS-BLITZ INTEGRATION SCRIPT")
    print(f"{'='*70}")
    print(f"Project  : {args.project}")
    print(f"API URL  : {internal_url}/api")
    print(f"MongoDB  : {os.getenv('DATABASE_URI', 'mongodb://localhost:27017')[:40]}...")
    if driver_ids:
        print(f"Drivers  : {', '.join(driver_ids)}")
    print(f"{'='*70}")

    try:
        integration = PMSBlitzIntegration()
        integration.sync_assigned_orders(args.project, driver_ids)
    except ValueError as e:
        print(f"\n❌ Configuration error: {e}")
        print("Please ensure BLITZ_USERNAME and BLITZ_PASSWORD environment variables are set")
        sys.exit(1)


if __name__ == "__main__":
    main()