import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ExcelChartGenerator:
    def __init__(self):
        self.primary_color = "1E3A8A"
        self.secondary_color = "3B82F6"
        self.success_color = "10B981"
        self.warning_color = "F59E0B"
        self.danger_color = "EF4444"
        self.light_bg = "F3F4F6"
        self.header_bg = "1E40AF"
        
    def create_workbook_with_charts(self, data, output_path):
        wb = openpyxl.Workbook()
        self.create_cover_sheet(wb, data)
        self.create_executive_summary(wb, data)
        self.create_performance_sheet(wb, data)
        self.create_insights_sheet(wb, data)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_path)
        return output_path
    
    def create_cover_sheet(self, wb, data):
        ws = wb.create_sheet("Dashboard Overview", 0)
        
        title_cell = ws.cell(row=2, column=2, value="PERFORMANCE ANALYTICS DASHBOARD")
        title_cell.font = Font(bold=True, size=24, color=self.primary_color)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:G2")
        
        subtitle_cell = ws.cell(row=3, column=2, value="Comprehensive Delivery Performance Report")
        subtitle_cell.font = Font(size=14, color="6B7280", italic=True)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B3:G3")
        
        ws.cell(row=5, column=2, value="Report Generated:").font = Font(bold=True, size=11)
        ws.cell(row=5, column=3, value=datetime.now().strftime("%d %B %Y, %H:%M")).font = Font(size=11)
        
        ws.cell(row=6, column=2, value="Report Period:").font = Font(bold=True, size=11)
        ws.cell(row=6, column=3, value="Current Month").font = Font(size=11)
        
        ws.cell(row=7, column=2, value="Department:").font = Font(bold=True, size=11)
        ws.cell(row=7, column=3, value="Operations & Logistics").font = Font(size=11)
        
        summary_data = data.get('summaryData', [])
        if summary_data:
            ws.cell(row=9, column=2, value="KEY METRICS OVERVIEW").font = Font(bold=True, size=14, color=self.primary_color)
            
            metric_row = 11
            for item in summary_data[:4]:
                metric_name = ws.cell(row=metric_row, column=2, value=item.get('Metric', ''))
                metric_name.font = Font(bold=True, size=11)
                metric_name.fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
                
                metric_value = ws.cell(row=metric_row, column=3, value=item.get('Value', ''))
                metric_value.font = Font(size=13, bold=True, color=self.secondary_color)
                metric_value.alignment = Alignment(horizontal="right")
                
                metric_unit = ws.cell(row=metric_row, column=4, value=item.get('Unit', ''))
                metric_unit.font = Font(size=10, color="6B7280")
                
                metric_row += 2
        
        ws.cell(row=20, column=2, value="REPORT SECTIONS").font = Font(bold=True, size=12, color=self.primary_color)
        
        sections = [
            ("Executive Summary", "High-level overview and key insights"),
            ("Performance Analysis", "Detailed performance metrics by location"),
            ("Strategic Insights", "Top performers and improvement areas")
        ]
        
        section_row = 22
        for section, desc in sections:
            ws.cell(row=section_row, column=2, value=f"• {section}").font = Font(bold=True, size=11)
            ws.cell(row=section_row, column=3, value=desc).font = Font(size=10, color="6B7280", italic=True)
            section_row += 1
        
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        ws.row_dimensions[2].height = 35
        ws.row_dimensions[3].height = 25
    
    def create_executive_summary(self, wb, data):
        ws = wb.create_sheet("Executive Summary")
        summary_data = data.get('summaryData', [])
        
        if not summary_data:
            return
        
        title = ws.cell(row=1, column=1, value="EXECUTIVE SUMMARY")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:E1")
        
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = Font(size=10, color="6B7280")
        
        headers = ["Metric", "Value", "Unit", "Status", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, item in enumerate(summary_data, 5):
            metric_cell = ws.cell(row=row, column=1, value=item.get('Metric', ''))
            metric_cell.font = Font(bold=True, size=11)
            metric_cell.fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            metric_cell.alignment = Alignment(horizontal="left", vertical="center")
            metric_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
            
            value_cell = ws.cell(row=row, column=2, value=item.get('Value', ''))
            value_cell.font = Font(size=12, bold=True)
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
            value_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
            
            if item.get('Unit') == 'percentage':
                try:
                    numeric_val = float(str(item.get('Value', '0')).replace('%', '').replace(',', ''))
                    value_cell.value = numeric_val
                    value_cell.number_format = '0.00"%"'
                    
                    if numeric_val >= 98:
                        value_cell.font = Font(size=12, bold=True, color=self.success_color)
                    elif numeric_val >= 95:
                        value_cell.font = Font(size=12, bold=True, color=self.warning_color)
                    else:
                        value_cell.font = Font(size=12, bold=True, color=self.danger_color)
                except:
                    pass
            
            unit_cell = ws.cell(row=row, column=3, value=item.get('Unit', ''))
            unit_cell.font = Font(size=10, color="6B7280")
            unit_cell.alignment = Alignment(horizontal="center", vertical="center")
            unit_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
            
            status = "Excellent"
            status_color = self.success_color
            if item.get('Unit') == 'percentage':
                try:
                    val = float(str(item.get('Value', '0')).replace('%', '').replace(',', ''))
                    if val < 95:
                        status = "Needs Attention"
                        status_color = self.danger_color
                    elif val < 98:
                        status = "Good"
                        status_color = self.warning_color
                except:
                    status = "N/A"
                    status_color = "6B7280"
            
            status_cell = ws.cell(row=row, column=4, value=status)
            status_cell.font = Font(bold=True, color="FFFFFF", size=10)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
            
            desc_cell = ws.cell(row=row, column=5, value=item.get('Description', ''))
            desc_cell.font = Font(size=10, color="374151")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="center")
            desc_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
        
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 45
        
        for i in range(4, 4 + len(summary_data) + 1):
            ws.row_dimensions[i].height = 25
        
        ws.freeze_panes = "A5"
        
        chart_start_row = len(summary_data) + 7
        self.create_summary_chart(ws, summary_data, chart_start_row)
    
    def create_summary_chart(self, ws, summary_data, start_row):
        ws.cell(row=start_row, column=1, value="PERFORMANCE METRICS VISUALIZATION").font = Font(bold=True, size=12, color=self.primary_color)
        
        chart_data_row = start_row + 2
        ws.cell(row=chart_data_row, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=chart_data_row, column=2, value="Value").font = Font(bold=True)
        
        numeric_data = []
        for item in summary_data:
            if item.get('Unit') in ['shipments', 'percentage']:
                try:
                    value_str = str(item.get('Value', '0')).replace(',', '').replace('%', '')
                    numeric_value = float(value_str)
                    numeric_data.append([item.get('Metric'), numeric_value])
                except:
                    continue
        
        for i, (metric, value) in enumerate(numeric_data, 1):
            ws.cell(row=chart_data_row + i, column=1, value=metric)
            ws.cell(row=chart_data_row + i, column=2, value=value)
        
        if numeric_data:
            chart = BarChart()
            chart.type = "col"
            chart.style = 11
            chart.title = "Key Metrics Overview"
            chart.y_axis.title = 'Value'
            chart.x_axis.title = 'Metrics'
            
            chart_data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(numeric_data))
            chart_categories = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(numeric_data))
            
            chart.add_data(chart_data, titles_from_data=True)
            chart.set_categories(chart_categories)
            chart.height = 12
            chart.width = 18
            
            ws.add_chart(chart, f"A{start_row + 2}")
    
    def create_performance_sheet(self, wb, data):
        ws = wb.create_sheet("Performance Analysis")
        performance_data = data.get('performanceData', [])
        
        if not performance_data:
            return
        
        sorted_performance = sorted(performance_data, key=lambda x: float(x.get('On Time Percentage', 0)), reverse=True)
        
        title = ws.cell(row=1, column=1, value="PERFORMANCE ANALYSIS BY LOCATION")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:H1")
        
        ws.cell(row=2, column=1, value=f"Total Locations: {len(sorted_performance)}").font = Font(size=10, color="6B7280")
        
        headers = ["Rank", "Location", "Category", "Total Shipments", "Late Shipments", "On Time %", "Late %", "Performance Level"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for idx, item in enumerate(sorted_performance, 1):
            row = idx + 4
            
            rank_cell = ws.cell(row=row, column=1, value=idx)
            rank_cell.alignment = Alignment(horizontal="center", vertical="center")
            rank_cell.font = Font(bold=True, size=10)
            
            if idx <= 3:
                rank_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            ws.cell(row=row, column=2, value=item.get('Short Name', ''))
            
            category_cell = ws.cell(row=row, column=3, value=item.get('Category', ''))
            category_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            total_cell = ws.cell(row=row, column=4, value=item.get('Total Shipments', 0))
            total_cell.number_format = '#,##0'
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            late_cell = ws.cell(row=row, column=5, value=item.get('Late Shipments', 0))
            late_cell.number_format = '#,##0'
            late_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            if item.get('Late Shipments', 0) > 0:
                late_cell.font = Font(color=self.danger_color, bold=True)
            
            ontime_cell = ws.cell(row=row, column=6, value=float(item.get('On Time Percentage', 0)))
            ontime_cell.number_format = '0.00"%"'
            ontime_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ontime_val = float(item.get('On Time Percentage', 0))
            if ontime_val >= 99.5:
                ontime_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                ontime_cell.font = Font(bold=True, color=self.success_color)
            elif ontime_val >= 98:
                ontime_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                ontime_cell.font = Font(bold=True, color=self.secondary_color)
            elif ontime_val >= 97:
                ontime_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                ontime_cell.font = Font(bold=True, color=self.warning_color)
            else:
                ontime_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                ontime_cell.font = Font(bold=True, color=self.danger_color)
            
            late_pct_cell = ws.cell(row=row, column=7, value=float(item.get('Late Percentage', 0)))
            late_pct_cell.number_format = '0.00"%"'
            late_pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            level = item.get('Performance Level', 'N/A')
            level_cell = ws.cell(row=row, column=8, value=level)
            level_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if level == 'Perfect':
                level_cell.fill = PatternFill(start_color=self.success_color, end_color=self.success_color, fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            elif level == 'Excellent':
                level_cell.fill = PatternFill(start_color=self.secondary_color, end_color=self.secondary_color, fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            elif level == 'Good':
                level_cell.fill = PatternFill(start_color=self.warning_color, end_color=self.warning_color, fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            else:
                level_cell.fill = PatternFill(start_color=self.danger_color, end_color=self.danger_color, fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB")
                )
        
        column_widths = [8, 30, 18, 16, 14, 12, 12, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        for i in range(4, 4 + len(sorted_performance) + 1):
            ws.row_dimensions[i].height = 22
        
        ws.freeze_panes = "C5"
        
        chart_row = len(sorted_performance) + 7
        self.create_performance_charts(ws, sorted_performance, chart_row)
    
    def create_performance_charts(self, ws, performance_data, start_row):
        ws.cell(row=start_row, column=1, value="PERFORMANCE VISUALIZATION").font = Font(bold=True, size=12, color=self.primary_color)
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 11
        chart1.title = "Top 15 Locations by On-Time Performance"
        chart1.y_axis.title = 'On Time Percentage (%)'
        chart1.x_axis.title = 'Location'
        
        data_range = Reference(ws, min_col=6, min_row=4, max_row=min(19, len(performance_data) + 4))
        categories = Reference(ws, min_col=2, min_row=5, max_row=min(19, len(performance_data) + 4))
        
        chart1.add_data(data_range, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.height = 13
        chart1.width = 20
        
        ws.add_chart(chart1, f"A{start_row + 2}")
        
        level_counts = {}
        for item in performance_data:
            level = item.get('Performance Level', 'N/A')
            level_counts[level] = level_counts.get(level, 0) + 1
        
        pie_start_row = start_row + 20
        ws.cell(row=pie_start_row, column=1, value="Performance Level Distribution").font = Font(bold=True, size=11)
        ws.cell(row=pie_start_row + 1, column=1, value="Level").font = Font(bold=True)
        ws.cell(row=pie_start_row + 1, column=2, value="Count").font = Font(bold=True)
        
        for i, (level, count) in enumerate(sorted(level_counts.items()), 1):
            ws.cell(row=pie_start_row + 1 + i, column=1, value=level)
            ws.cell(row=pie_start_row + 1 + i, column=2, value=count)
        
        pie_chart = PieChart()
        pie_chart.title = "Distribution by Performance Level"
        
        pie_data = Reference(ws, min_col=2, min_row=pie_start_row + 1, max_row=pie_start_row + 1 + len(level_counts))
        pie_categories = Reference(ws, min_col=1, min_row=pie_start_row + 2, max_row=pie_start_row + 1 + len(level_counts))
        
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        pie_chart.height = 12
        pie_chart.width = 16
        
        ws.add_chart(pie_chart, f"K{start_row + 2}")
    
    def create_insights_sheet(self, wb, data):
        ws = wb.create_sheet("Strategic Insights")
        insights_data = data.get('insightsData', [])
        
        if not insights_data:
            return
        
        title = ws.cell(row=1, column=1, value="STRATEGIC INSIGHTS & RECOMMENDATIONS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:G1")
        
        headers = ["Category", "Location", "Short Name", "Volume", "On Time %", "Late %", "Performance"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.border = Border(bottom=Side(style="medium", color=self.primary_color))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        top_performers = sorted([item for item in insights_data if item.get('Category') == 'Top Performer'], 
                               key=lambda x: float(x.get('Percentage', 0)), reverse=True)
        priority_areas = sorted([item for item in insights_data if item.get('Category') == 'Priority Area'], 
                               key=lambda x: float(x.get('Percentage', 0)))
        volume_leaders = sorted([item for item in insights_data if item.get('Category') == 'Volume Leader'], 
                               key=lambda x: int(x.get('Value', 0)), reverse=True)
        
        current_row = 4
        
        categories_config = [
            ("TOP PERFORMERS", top_performers, "D1FAE5", self.success_color),
            ("PRIORITY AREAS (NEEDS IMPROVEMENT)", priority_areas, "FEE2E2", self.danger_color),
            ("VOLUME LEADERS", volume_leaders, "DBEAFE", self.secondary_color)
        ]
        
        for category_name, category_data, bg_color, text_color in categories_config:
            if not category_data:
                continue
            
            category_cell = ws.cell(row=current_row, column=1, value=category_name)
            category_cell.font = Font(bold=True, size=12, color=text_color)
            category_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            category_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"A{current_row}:G{current_row}")
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            for item in category_data:
                ws.cell(row=current_row, column=1, value=item.get('Category', ''))
                ws.cell(row=current_row, column=2, value=item.get('Location', ''))
                ws.cell(row=current_row, column=3, value=item.get('Short Name', ''))
                
                volume_cell = ws.cell(row=current_row, column=4, value=item.get('Value', 0))
                volume_cell.number_format = '#,##0'
                volume_cell.alignment = Alignment(horizontal="right", vertical="center")
                
                percentage_cell = ws.cell(row=current_row, column=5, value=float(item.get('Percentage', 0)))
                percentage_cell.number_format = '0.00"%"'
                percentage_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                late_pct = 100 - float(item.get('Percentage', 0))
                late_cell = ws.cell(row=current_row, column=6, value=late_pct)
                late_cell.number_format = '0.00"%"'
                late_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                perf_cell = ws.cell(row=current_row, column=7, value=item.get('Performance Level', ''))
                perf_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = Border(
                        left=Side(style="thin", color="D1D5DB"),
                        right=Side(style="thin", color="D1D5DB"),
                        top=Side(style="thin", color="D1D5DB"),
                        bottom=Side(style="thin", color="D1D5DB")
                    )
                
                current_row += 1
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        
        ws.freeze_panes = "A4"
        
        if top_performers:
            chart_row = current_row + 2
            self.create_insights_chart(ws, top_performers, chart_row)
    
    def create_insights_chart(self, ws, top_performers, start_row):
        ws.cell(row=start_row, column=1, value="TOP PERFORMERS COMPARISON").font = Font(bold=True, size=12, color=self.primary_color)
        
        chart_data_row = start_row + 2
        ws.cell(row=chart_data_row, column=1, value="Location").font = Font(bold=True)
        ws.cell(row=chart_data_row, column=2, value="Performance (%)").font = Font(bold=True)
        
        for i, item in enumerate(top_performers[:10], 1):
            ws.cell(row=chart_data_row + i, column=1, value=item.get('Short Name', ''))
            ws.cell(row=chart_data_row + i, column=2, value=float(item.get('Percentage', 0)))
        
        chart = BarChart()
        chart.type = "bar"
        chart.style = 13
        chart.title = "Top 10 Performers"
        chart.x_axis.title = 'On-Time Performance (%)'
        chart.y_axis.title = 'Location'
        
        chart_data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + min(len(top_performers), 10))
        chart_categories = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + min(len(top_performers), 10))
        
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_categories)
        chart.height = 12
        chart.width = 18
        
        ws.add_chart(chart, f"D{start_row}")

def main():
    try:
        if len(sys.argv) != 3:
            raise ValueError("Usage: python chart_generator.py <input_json_path> <output_excel_path>")
        
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        generator = ExcelChartGenerator()
        result_path = generator.create_workbook_with_charts(data, output_path)
        
        print(json.dumps({
            "success": True,
            "output_path": result_path,
            "message": "Professional dashboard created successfully"
        }))
    
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()