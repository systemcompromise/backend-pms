import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, PieChart, RadarChart, AreaChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MitraPerformanceChartGeneratorFormula:
    def __init__(self):
        self.primary_color = "1E3A8A"
        self.secondary_color = "3B82F6"
        self.success_color = "10B981"
        self.warning_color = "F59E0B"
        self.danger_color = "EF4444"
        self.light_bg = "F3F4F6"
        self.header_bg = "1E40AF"
        
    def create_workbook_with_charts(self, data, output_path):
        wb = openpyxl.Workbook()
        
        data_quality = data.get('dataQuality', {})
        has_valid_trends = data_quality.get('hasValidTrends', False)
        period_type = data.get('periodType', 'monthly')
        applied_filters = data.get('appliedFilters', {})
        
        self.create_constants_sheet(wb)
        self.create_data_quality_warning_sheet(wb, data_quality, data)
        self.create_shipment_data_sheet(wb, data, period_type)
        self.create_executive_summary_sheet(wb, data, has_valid_trends, period_type)
        self.create_performance_metrics_with_formulas(wb, data, period_type)
        self.create_cost_analysis_dashboard(wb, data, period_type)
        
        if has_valid_trends:
            self.create_trend_analysis_with_formulas(wb, data, period_type)
        else:
            self.create_limited_trend_sheet(wb, data, period_type)
        
        self.create_project_analysis_with_formulas(wb, data, period_type)
        self.create_operational_insights_dashboard(wb, data, period_type)
        self.create_performance_overview_sheet(wb, data, has_valid_trends, period_type)
        self.create_visual_dashboard(wb, data, has_valid_trends, period_type)
        
        if has_valid_trends:
            self.create_advanced_analytics_dashboard(wb, data, period_type)
        
        self.create_management_kpi_dashboard(wb, data, has_valid_trends, period_type)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.active = wb['Executive Summary'] if has_valid_trends else wb['Data Quality Warning']
        
        wb.save(output_path)
        return output_path
    
    def safe_float(self, value, default=0.0):
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def safe_int(self, value, default=0):
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default
    
    def create_constants_sheet(self, wb):
        ws = wb.create_sheet("Constants", 0)
        ws.sheet_state = 'hidden'
        
        title = ws.cell(row=1, column=1, value="SYSTEM CONSTANTS")
        title.font = Font(bold=True, size=14, color=self.primary_color)
        
        ws.cell(row=3, column=1, value="Constant Name").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Value").font = Font(bold=True)
        ws.cell(row=3, column=3, value="Description").font = Font(bold=True)
        
        constants = [
            ("DELIVERY_RATE_TARGET", 95, "Target delivery success rate (95%)"),
            ("ONTIME_RATE_TARGET", 90, "Target on-time delivery rate (90%)"),
            ("ACTIVITY_BASELINE", 100, "Baseline for activity level calculation"),
            ("WEIGHT_DELIVERY_RATE", 0.30, "Weight for delivery rate in score (30%)"),
            ("WEIGHT_ONTIME_RATE", 0.25, "Weight for on-time rate in score (25%)"),
            ("WEIGHT_ACTIVITY", 0.20, "Weight for activity level in score (20%)"),
            ("WEIGHT_CONSISTENCY", 0.15, "Weight for consistency in score (15%)"),
            ("WEIGHT_GROWTH", 0.10, "Weight for growth in score (10%)"),
            ("GROWTH_BASELINE", 50, "Baseline score for growth calculation (50)"),
            ("MAX_SCORE", 100, "Maximum performance score (100)"),
            ("CANCEL_PENALTY_MULTIPLIER", 10, "Multiplier for cancellation penalty")
        ]
        
        for idx, (name, value, desc) in enumerate(constants, 4):
            ws.cell(row=idx, column=1, value=name).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
            ws.cell(row=idx, column=3, value=desc).font = Font(italic=True, size=9)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
    
    def create_data_quality_warning_sheet(self, wb, data_quality, full_data):
        ws = wb.create_sheet("Data Quality Warning", 1)
        
        has_valid = data_quality.get('hasValidTrends', False)
        trend_count = data_quality.get('trendCount', 0)
        shipment_count = data_quality.get('shipmentCount', 0)
        period_type = full_data.get('periodType', 'monthly')
        
        if has_valid:
            title = ws.cell(row=2, column=2, value="✅ DATA QUALITY: COMPLETE ANALYSIS")
            title.font = Font(bold=True, size=20, color=self.success_color)
        else:
            title = ws.cell(row=2, column=2, value="⚠️ DATA QUALITY: LIMITED ANALYSIS")
            title.font = Font(bold=True, size=20, color=self.warning_color)
        
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:H2")
        ws.row_dimensions[2].height = 40
        
        ws.cell(row=4, column=2, value="CURRENT DATA STATUS").font = Font(bold=True, size=14, color=self.primary_color)
        
        ws.cell(row=6, column=2, value="Analysis Period Type").font = Font(bold=True, size=11)
        ws.cell(row=6, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
        ws.cell(row=6, column=4, value=period_type.capitalize()).font = Font(size=12, bold=True, color=self.success_color if has_valid else self.warning_color)
        ws.cell(row=6, column=4).alignment = Alignment(horizontal="center")
        
        ws.cell(row=7, column=2, value="Delivery Periods Available").font = Font(bold=True, size=11)
        ws.cell(row=7, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
        ws.cell(row=7, column=4, value=trend_count).font = Font(size=12, bold=True, color=self.success_color if has_valid else self.warning_color)
        ws.cell(row=7, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=7, column=5, value="periods").font = Font(size=10, italic=True)
        
        ws.cell(row=8, column=2, value="Total Shipment Records").font = Font(bold=True, size=11)
        ws.cell(row=8, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
        ws.cell(row=8, column=4, value=shipment_count).font = Font(size=12, bold=True, color=self.success_color if has_valid else self.warning_color)
        ws.cell(row=8, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=8, column=5, value="records").font = Font(size=10, italic=True)
        
        ws.cell(row=9, column=2, value="Minimum Required Periods").font = Font(bold=True, size=11)
        ws.cell(row=9, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
        ws.cell(row=9, column=4, value=2).font = Font(size=12, bold=True, color=self.success_color if has_valid else self.warning_color)
        ws.cell(row=9, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=9, column=5, value="periods").font = Font(size=10, italic=True)
        
        ws.cell(row=10, column=2, value="Analysis Status").font = Font(bold=True, size=11)
        ws.cell(row=10, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
        ws.cell(row=10, column=4, value="COMPLETE" if has_valid else "LIMITED").font = Font(size=12, bold=True, color=self.success_color if has_valid else self.warning_color)
        ws.cell(row=10, column=4).alignment = Alignment(horizontal="center")
        
        ws.cell(row=12, column=2, value="BUSINESS IMPACT ASSESSMENT").font = Font(bold=True, size=12, color=self.primary_color)
        
        if has_valid:
            impact_assessment = [
                ("✅ Strategic Planning", "Multi-period trends enable accurate forecasting and resource planning", True),
                ("✅ Performance Evaluation", "Complete trend analysis supports comprehensive performance reviews", True),
                ("✅ Cost Optimization", "Historical data enables identification of cost patterns and optimization opportunities", True),
                ("✅ Risk Management", "Trend analysis helps identify potential operational risks and mitigation strategies", True),
                ("✅ Growth Analysis", "Period-over-period comparisons support accurate growth assessment", True)
            ]
        else:
            impact_assessment = [
                ("❌ Strategic Planning", "Limited to single period - insufficient for forecasting and resource planning", False),
                ("❌ Performance Evaluation", "Unable to assess performance trends or identify improvement patterns", False),
                ("❌ Cost Optimization", "Cannot identify cost patterns or optimization opportunities across periods", False),
                ("❌ Risk Management", "Limited ability to identify operational risks or trends", False),
                ("❌ Growth Analysis", "Cannot assess period-over-period growth or performance changes", False)
            ]
        
        impact_row = 14
        for feature, description, available in impact_assessment:
            ws.cell(row=impact_row, column=2, value=feature).font = Font(bold=True, size=10, color=self.success_color if available else self.danger_color)
            ws.merge_cells(f"B{impact_row}:C{impact_row}")
            
            ws.cell(row=impact_row, column=4, value=description).font = Font(size=9, italic=True)
            ws.merge_cells(f"D{impact_row}:H{impact_row}")
            
            impact_row += 1
        
        if not has_valid:
            ws.cell(row=impact_row + 2, column=2, value="📋 RECOMMENDATIONS FOR FULL ANALYSIS").font = Font(bold=True, size=12, color=self.primary_color)
            
            recommendations = [
                "1. Ensure delivery data spans at least 2 different periods for trend analysis",
                "2. Each period should have sufficient delivery records for statistical significance",
                "3. Delivery dates must be properly formatted (DD/MM/YYYY) for accurate period grouping",
                "4. Re-export the report after adding more delivery data to unlock full analytical capabilities",
                "5. Consider implementing regular data collection processes to maintain analysis quality"
            ]
            
            rec_row = impact_row + 4
            for recommendation in recommendations:
                ws.cell(row=rec_row, column=2, value=recommendation).font = Font(size=10)
                ws.merge_cells(f"B{rec_row}:H{rec_row}")
                rec_row += 1
        
        footer_row = impact_row + 20
        ws.cell(row=footer_row, column=2, value=f"Report Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}").font = Font(size=9, italic=True, color="6B7280")
        ws.cell(row=footer_row + 1, column=2, value="All calculations use Excel formulas for transparency and auditability").font = Font(size=9, italic=True, color="6B7280")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 20
    
    def create_limited_trend_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Delivery Trends")
        trends = data.get('trends', [])
        
        title = ws.cell(row=1, column=1, value=f"⚠️ LIMITED TREND DATA - {period_type.upper()} STATISTICS ONLY")
        title.font = Font(bold=True, size=16, color=self.warning_color)
        ws.merge_cells("A1:G1")
        
        ws.cell(row=3, column=1, value=f"Available Periods: {len(trends)}").font = Font(size=12, bold=True)
        ws.cell(row=4, column=1, value="Minimum Required: 2 periods for trend analysis").font = Font(size=11, italic=True)
        
        ws.cell(row=6, column=1, value="AVAILABLE DATA").font = Font(bold=True, size=12, color=self.primary_color)
        
        headers = ["Period", "Deliveries", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for idx, trend in enumerate(trends, 9):
            ws.cell(row=idx, column=1, value=trend.get('month', 'Unknown'))
            ws.cell(row=idx, column=2, value=int(self.safe_float(trend.get('deliveries', 0)))).number_format = '#,##0'
            ws.cell(row=idx, column=3, value="Baseline period - no comparison available")
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        return "Delivery Trends"
    
    def format_date(self, date_str):
        if not date_str or date_str == 'N/A':
            return 'N/A'
        try:
            date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return date_obj.strftime('%d %B %Y')
        except:
            return date_str
    
    def create_executive_summary_sheet(self, wb, data, has_valid_trends, period_type):
        ws = wb.create_sheet("Executive Summary")
        profile = data.get('profile', {})
        metrics = data.get('metrics', {})
        
        if has_valid_trends:
            title = ws.cell(row=2, column=2, value=f"EXECUTIVE SUMMARY - {period_type.upper()} PERFORMANCE ANALYSIS")
            title.font = Font(bold=True, size=20, color=self.primary_color)
        else:
            title = ws.cell(row=2, column=2, value=f"EXECUTIVE SUMMARY (LIMITED DATA) - {period_type.upper()} PERFORMANCE")
            title.font = Font(bold=True, size=20, color=self.warning_color)
        
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B2:G2")
        ws.row_dimensions[2].height = 35
        
        subtitle_text = "Strategic Performance Metrics for Management Decision-Making" if has_valid_trends else "Basic Performance Metrics - Add More Data for Strategic Analysis"
        subtitle = ws.cell(row=3, column=2, value=subtitle_text)
        subtitle.font = Font(size=12, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("B3:G3")
        
        ws.cell(row=5, column=2, value="MITRA PROFILE").font = Font(bold=True, size=12, color=self.primary_color)
        
        profile_data = [
            ("Name:", profile.get('name', 'N/A')),
            ("Driver ID:", profile.get('driverId', 'N/A')),
            ("Phone:", profile.get('phone', 'N/A')),
            ("City:", profile.get('city', 'N/A')),
            ("Status:", profile.get('status', 'N/A')),
            ("Joined:", self.format_date(profile.get('joinedDate', 'N/A')))
        ]
        
        row = 7
        for label, value in profile_data:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=10)
            ws.cell(row=row, column=3, value=value).font = Font(size=10)
            row += 1
        
        ws.cell(row=13, column=2, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=12, color=self.primary_color)
        ws.cell(row=14, column=2, value="(Strategic metrics for business decision-making)").font = Font(size=9, italic=True, color="6B7280")
        
        kpi_row = 16
        kpi_labels = [
            "Total Deliveries",
            "Delivery Success Rate",
            "On-Time Delivery Rate",
            "Average Distance",
            "Total Cost",
            "Average Cost per Delivery",
            "Cost per Kilometer",
            "Cancellation Rate",
            "Growth Rate",
            "Unique Projects",
            "Unique Hubs"
        ]
        
        kpi_formulas = [
            "=IFERROR('Performance Metrics'!C4,0)",
            "=IFERROR('Performance Metrics'!C10*100,0)",
            "=IFERROR('Performance Metrics'!C6*100,0)",
            "=IFERROR('Performance Metrics'!C7,0)",
            "=IFERROR('Cost Analysis'!C6,0)",
            "=IFERROR('Cost Analysis'!C7,0)",
            "=IFERROR('Cost Analysis'!C8,0)",
            "=IFERROR('Performance Metrics'!C11*100,0)",
            "=IFERROR('Performance Metrics'!C12*100,0)",
            "=IFERROR('Performance Metrics'!C8,0)",
            "=IFERROR('Performance Metrics'!C9,0)"
        ]
        
        kpi_formats = [
            '#,##0',
            '0.00"%"',
            '0.00"%"',
            '0.00" km"',
            'Rp #,##0',
            'Rp #,##0',
            'Rp #,##0"/km"',
            '0.00"%"',
            '+0.00"%";-0.00"%"',
            '#,##0',
            '#,##0'
        ]
        
        for label, formula, number_format in zip(kpi_labels, kpi_formulas, kpi_formats):
            ws.cell(row=kpi_row, column=2, value=label).font = Font(bold=True, size=10)
            ws.cell(row=kpi_row, column=2).fill = PatternFill(start_color=self.light_bg, end_color=self.light_bg, fill_type="solid")
            
            value_cell = ws.cell(row=kpi_row, column=3, value=formula)
            value_cell.font = Font(size=12, bold=True, color=self.secondary_color)
            value_cell.alignment = Alignment(horizontal="center")
            value_cell.number_format = number_format
            
            kpi_row += 1
        
        score_row = kpi_row + 2
        ws.cell(row=score_row, column=2, value="OVERALL PERFORMANCE SCORE").font = Font(bold=True, size=12, color=self.primary_color)
        ws.cell(row=score_row + 1, column=2, value="(Strategic performance assessment)").font = Font(size=9, italic=True, color="6B7280")
        score_cell = ws.cell(row=score_row + 2, column=3, value="=IFERROR('Performance Metrics'!D21,0)")
        score_cell.font = Font(size=24, bold=True, color=self.success_color)
        score_cell.number_format = '0.00'
        score_cell.alignment = Alignment(horizontal="center")
        
        ws.cell(row=score_row + 4, column=2, value="PERFORMANCE CATEGORY").font = Font(bold=True, size=11, color=self.primary_color)
        category_cell = ws.cell(row=score_row + 5, column=3, value="=IF(D23>=90,\"Excellent\",IF(D23>=80,\"Very Good\",IF(D23>=70,\"Good\",\"Fair\")))")
        category_cell.font = Font(size=14, bold=True)
        category_cell.alignment = Alignment(horizontal="center")
        
        if not has_valid_trends:
            ws.cell(row=score_row + 7, column=2, value="⚠️ LIMITED ANALYSIS").font = Font(bold=True, size=11, color=self.warning_color)
            ws.cell(row=score_row + 8, column=2, value="Strategic metrics unavailable (requires 2+ periods)").font = Font(size=10, italic=True)
            ws.merge_cells(f"B{score_row + 8}:G{score_row + 8}")
        
        ws.cell(row=score_row + 10, column=2, value="BUSINESS RECOMMENDATIONS").font = Font(bold=True, size=11, color=self.primary_color)
        
        if has_valid_trends:
            recommendations = [
                "• Maintain current performance levels if score ≥ 80",
                "• Focus on improving on-time delivery rate if < 90%",
                "• Optimize cost per delivery if above industry average",
                "• Leverage growth trends for expansion planning",
                "• Address cancellation patterns if rate > 5%"
            ]
        else:
            recommendations = [
                "• Add more delivery periods for comprehensive analysis",
                "• Establish consistent data collection processes",
                "• Implement performance tracking across multiple periods",
                "• Develop trend analysis capabilities for strategic planning"
            ]
        
        rec_row = score_row + 12
        for recommendation in recommendations:
            ws.cell(row=rec_row, column=2, value=recommendation).font = Font(size=10)
            ws.merge_cells(f"B{rec_row}:G{rec_row}")
            rec_row += 1
        
        ws.cell(row=rec_row + 2, column=2, value="REPORT INFORMATION").font = Font(bold=True, size=11, color=self.primary_color)
        ws.cell(row=rec_row + 3, column=2, value="Generated:").font = Font(bold=True)
        ws.cell(row=rec_row + 3, column=3, value=datetime.now().strftime("%d %B %Y, %H:%M"))
        ws.cell(row=rec_row + 4, column=2, value="Analysis Type:").font = Font(bold=True)
        ws.cell(row=rec_row + 4, column=3, value=f"{period_type.capitalize()} Strategic Analysis" if has_valid_trends else f"Limited {period_type.capitalize()} Analysis")
        ws.cell(row=rec_row + 5, column=2, value="Calculation Method:").font = Font(bold=True)
        ws.cell(row=rec_row + 5, column=3, value="Excel formulas for transparency")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
    
    def clean_string(self, value):
        if value is None or value == '' or value == '-':
            return '-'
        return str(value).strip()
    
    def clean_number(self, value):
        if value is None or value == '' or value == '-':
            return 0
        try:
            return float(str(value).replace(',', '.'))
        except:
            return 0
    
    def is_on_time(self, sla_value):
        if not sla_value or sla_value == '-':
            return 0
        sla_lower = str(sla_value).lower()
        return 1 if 'on time' in sla_lower or 'ontime' in sla_lower else 0
    
    def extract_period_info(self, date_str, period_type):
        if not date_str or date_str == '-':
            return None, None, None, None
        
        try:
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                day_num = int(day)
                month_num = int(month)
                year_num = int(year)
                
                month_names = ["", "January", "February", "March", "April", "May", "June", 
                              "July", "August", "September", "October", "November", "December"]
                
                if period_type == 'daily':
                    display = date_str
                    return display, month_num, year_num, f"{year_num}{str(month_num).zfill(2)}{str(day_num).zfill(2)}"
                    
                elif period_type == 'weekly':
                    date_obj = datetime(year_num, month_num, day_num)
                    week_num = date_obj.isocalendar()[1]
                    display = f"Week {week_num} - {month_names[month_num]} {year_num}"
                    return display, month_num, year_num, f"{year_num}{str(month_num).zfill(2)}{str(week_num).zfill(2)}"
                    
                elif period_type == 'monthly':
                    display = f"{month_names[month_num]} {year_num}"
                    return display, month_num, year_num, f"{year_num}{str(month_num).zfill(2)}"
                    
                elif period_type == 'yearly':
                    display = str(year_num)
                    return display, None, year_num, str(year_num)
        except:
            pass
        
        return None, None, None, None
    
    def create_shipment_data_sheet(self, wb, data, period_type):
        ws = wb.create_sheet("Shipment Data", 2)
        shipment_data = data.get('shipmentData', [])
        
        if len(shipment_data) == 0:
            ws.cell(row=1, column=1, value="No shipment data available").font = Font(bold=True, color="FF0000")
            return
        
        title = ws.cell(row=1, column=1, value=f"RAW SHIPMENT DATA - {period_type.upper()} FILTERED (CHRONOLOGICAL)")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:W1")
        
        ws.cell(row=2, column=1, value="Sorted: Oldest deliveries first | All calculations reference this data").font = Font(size=9, italic=True, color="6B7280")
        ws.merge_cells("A2:W2")
        
        headers = ["Client Name", "Project Name", "Delivery Date", "Drop Point", "Hub", 
                   "Order Code", "Weight", "Distance (km)", "Mitra Code", "Mitra Name", 
                   "Receiving Date", "Vehicle Type", "Cost", "SLA", "Weekly", "Is On-Time",
                   "Display Period", "Month", "Year", "Cost Numeric", "Distance Numeric", 
                   "Month Num", "Year Num", "Sort Key"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        processed_shipments = []
        for shipment in shipment_data:
            date_str = self.clean_string(shipment.get('delivery_date'))
            display_period, month, year, sort_key = self.extract_period_info(date_str, period_type)
            
            processed_shipments.append({
                'data': shipment,
                'display_period': display_period if display_period else date_str,
                'month': month,
                'year': year,
                'sort_key': sort_key if sort_key else '99999999',
                'date_str': date_str
            })
        
        processed_shipments.sort(key=lambda x: x['sort_key'])
        
        for row_idx, item in enumerate(processed_shipments, 4):
            shipment = item['data']
            
            ws.cell(row=row_idx, column=1, value=self.clean_string(shipment.get('client_name')))
            ws.cell(row=row_idx, column=2, value=self.clean_string(shipment.get('project_name')))
            ws.cell(row=row_idx, column=3, value=item['date_str'])
            ws.cell(row=row_idx, column=4, value=self.clean_string(shipment.get('drop_point')))
            ws.cell(row=row_idx, column=5, value=self.clean_string(shipment.get('hub')))
            ws.cell(row=row_idx, column=6, value=self.clean_string(shipment.get('order_code')))
            ws.cell(row=row_idx, column=7, value=self.clean_string(shipment.get('weight')))
            
            distance = self.clean_number(shipment.get('distance_km'))
            ws.cell(row=row_idx, column=8, value=distance).number_format = '0.00'
            
            ws.cell(row=row_idx, column=9, value=self.clean_string(shipment.get('mitra_code')))
            ws.cell(row=row_idx, column=10, value=self.clean_string(shipment.get('mitra_name')))
            ws.cell(row=row_idx, column=11, value=self.clean_string(shipment.get('receiving_date')))
            ws.cell(row=row_idx, column=12, value=self.clean_string(shipment.get('vehicle_type')))
            
            cost_str = self.clean_string(shipment.get('cost'))
            ws.cell(row=row_idx, column=13, value=cost_str)
            
            sla_value = self.clean_string(shipment.get('sla'))
            ws.cell(row=row_idx, column=14, value=sla_value)
            ws.cell(row=row_idx, column=15, value=self.clean_string(shipment.get('weekly')))
            
            is_on_time = self.is_on_time(sla_value)
            ws.cell(row=row_idx, column=16, value=is_on_time)
            
            ws.cell(row=row_idx, column=17, value=item['display_period'])
            ws.cell(row=row_idx, column=18, value=item['month'] if item['month'] else '')
            ws.cell(row=row_idx, column=19, value=item['year'] if item['year'] else '')
            
            cost_numeric = self.clean_number(cost_str)
            ws.cell(row=row_idx, column=20, value=cost_numeric).number_format = '#,##0'
            ws.cell(row=row_idx, column=21, value=distance).number_format = '0.00'
            ws.cell(row=row_idx, column=22, value=item['month'] if item['month'] else '')
            ws.cell(row=row_idx, column=23, value=item['year'] if item['year'] else '')
            ws.cell(row=row_idx, column=24, value=item['sort_key'])
        
        for col in range(1, 25):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_performance_metrics_with_formulas(self, wb, data, period_type):
        ws = wb.create_sheet("Performance Metrics")
        profile = data.get('profile', {})
        metrics = data.get('metrics', {})
        shipment_count = len(data.get('shipmentData', []))
        last_data_row = 3 + shipment_count
        
        title = ws.cell(row=1, column=1, value=f"PERFORMANCE METRICS - {period_type.upper()} EXCEL FORMULAS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        ws.cell(row=2, column=1, value="All calculations use formulas referencing Shipment Data and Constants sheets").font = Font(size=9, italic=True, color="6B7280")
        ws.merge_cells("A2:F2")
        
        headers = ["Metric", "Formula", "Result", "Unit"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if shipment_count > 0:
            total_shipments_row = 4
            ws.cell(row=total_shipments_row, column=1, value="Total Deliveries").font = Font(bold=True)
            ws.cell(row=total_shipments_row, column=2, value=f"=COUNTA('Shipment Data'!A4:A{last_data_row})").font = Font(italic=True, size=9)
            ws.cell(row=total_shipments_row, column=3, value=f"=IFERROR(B{total_shipments_row},0)").number_format = '#,##0'
            ws.cell(row=total_shipments_row, column=4, value="deliveries")
            
            on_time_row = 5
            ws.cell(row=on_time_row, column=1, value="On-Time Deliveries").font = Font(bold=True)
            ws.cell(row=on_time_row, column=2, value=f"=SUM('Shipment Data'!P4:P{last_data_row})").font = Font(italic=True, size=9)
            ws.cell(row=on_time_row, column=3, value=f"=IFERROR(B{on_time_row},0)").number_format = '#,##0'
            ws.cell(row=on_time_row, column=4, value="deliveries")
            
            on_time_rate_row = 6
            ws.cell(row=on_time_rate_row, column=1, value="On-Time Rate").font = Font(bold=True)
            ws.cell(row=on_time_rate_row, column=2, value=f"=IFERROR(IF(C{total_shipments_row}=0,0,C{on_time_row}/C{total_shipments_row}),0)").font = Font(italic=True, size=9)
            ws.cell(row=on_time_rate_row, column=3, value=f"=IFERROR(B{on_time_rate_row},0)").number_format = '0.00'
            ws.cell(row=on_time_rate_row, column=4, value="percentage")
            
            avg_distance_row = 7
            ws.cell(row=avg_distance_row, column=1, value="Average Distance").font = Font(bold=True)
            ws.cell(row=avg_distance_row, column=2, value=f"=IFERROR(IF(COUNTIF('Shipment Data'!U4:U{last_data_row},\">0\")=0,0,SUM('Shipment Data'!U4:U{last_data_row})/COUNTIF('Shipment Data'!U4:U{last_data_row},\">0\")),0)").font = Font(italic=True, size=9)
            ws.cell(row=avg_distance_row, column=3, value=f"=IFERROR(B{avg_distance_row},0)").number_format = '0.00'
            ws.cell(row=avg_distance_row, column=4, value="km")
        else:
            ws.cell(row=4, column=1, value="No shipment data available").font = Font(color="FF0000")
            return
        
        unique_projects_row = 8
        ws.cell(row=unique_projects_row, column=1, value="Unique Projects").font = Font(bold=True)
        ws.cell(row=unique_projects_row, column=2, value=f"=IFERROR(SUMPRODUCT(1/COUNTIF('Shipment Data'!B4:B{last_data_row},'Shipment Data'!B4:B{last_data_row}&\"\")),0)").font = Font(italic=True, size=9)
        ws.cell(row=unique_projects_row, column=3, value=f"=IFERROR(B{unique_projects_row},0)").number_format = '#,##0'
        ws.cell(row=unique_projects_row, column=4, value="projects")
        
        unique_hubs_row = 9
        ws.cell(row=unique_hubs_row, column=1, value="Unique Hubs").font = Font(bold=True)
        ws.cell(row=unique_hubs_row, column=2, value=f"=IFERROR(SUMPRODUCT(1/COUNTIF('Shipment Data'!E4:E{last_data_row},'Shipment Data'!E4:E{last_data_row}&\"\")),0)").font = Font(italic=True, size=9)
        ws.cell(row=unique_hubs_row, column=3, value=f"=IFERROR(B{unique_hubs_row},0)").number_format = '#,##0'
        ws.cell(row=unique_hubs_row, column=4, value="hubs")
        
        delivery_rate_row = 10
        ws.cell(row=delivery_rate_row, column=1, value="Delivery Success Rate").font = Font(bold=True)
        ws.cell(row=delivery_rate_row, column=2, value=f"=IFERROR(IF(C{total_shipments_row}=0,0,(C{total_shipments_row}-({self.safe_float(metrics.get('cancelRate', 0))/100}*C{total_shipments_row}))/C{total_shipments_row}),0)").font = Font(italic=True, size=9)
        ws.cell(row=delivery_rate_row, column=3, value=f"=IFERROR(B{delivery_rate_row},0)").number_format = '0.00'
        ws.cell(row=delivery_rate_row, column=4, value="percentage")
        
        cancel_rate_row = 11
        ws.cell(row=cancel_rate_row, column=1, value="Cancellation Rate").font = Font(bold=True)
        ws.cell(row=cancel_rate_row, column=2, value=f"=IFERROR(1-C{delivery_rate_row},0)").font = Font(italic=True, size=9)
        ws.cell(row=cancel_rate_row, column=3, value=f"=IFERROR(B{cancel_rate_row},0)").number_format = '0.00'
        ws.cell(row=cancel_rate_row, column=4, value="percentage")
        
        growth_rate_row = 12
        if period_type == 'daily':
            ws.cell(row=growth_rate_row, column=1, value="Growth Rate").font = Font(bold=True)
            ws.cell(row=growth_rate_row, column=2, value="=IFERROR(IF(COUNTA('Shipment Data'!Q4:Q1000)<=1,0,('Shipment Data'!Q5-INDEX('Shipment Data'!Q:Q,COUNTA('Shipment Data'!Q4:Q1000)+4))/INDEX('Shipment Data'!Q:Q,COUNTA('Shipment Data'!Q4:Q1000)+4)),0)").font = Font(italic=True, size=9)
        elif period_type == 'weekly':
            ws.cell(row=growth_rate_row, column=1, value="Growth Rate").font = Font(bold=True)
            ws.cell(row=growth_rate_row, column=2, value="=IFERROR(IF(COUNTA('Shipment Data'!Q4:Q1000)<=1,0,('Shipment Data'!Q5-INDEX('Shipment Data'!Q:Q,COUNTA('Shipment Data'!Q4:Q1000)+4))/INDEX('Shipment Data'!Q:Q,COUNTA('Shipment Data'!Q4:Q1000)+4)),0)").font = Font(italic=True, size=9)
        elif period_type == 'monthly':
            ws.cell(row=growth_rate_row, column=1, value="Growth Rate").font = Font(bold=True)
            ws.cell(row=growth_rate_row, column=2, value="=IFERROR(IF(COUNTA('Shipment Data'!R4:R1000)<=1,0,('Shipment Data'!R5-INDEX('Shipment Data'!R:R,COUNTA('Shipment Data'!R4:R1000)+4))/INDEX('Shipment Data'!R:R,COUNTA('Shipment Data'!R4:R1000)+4)),0)").font = Font(italic=True, size=9)
        else:  # yearly
            ws.cell(row=growth_rate_row, column=1, value="Growth Rate").font = Font(bold=True)
            ws.cell(row=growth_rate_row, column=2, value="=IFERROR(IF(COUNTA('Shipment Data'!S4:S1000)<=1,0,('Shipment Data'!S5-INDEX('Shipment Data'!S:S,COUNTA('Shipment Data'!S4:S1000)+4))/INDEX('Shipment Data'!S:S,COUNTA('Shipment Data'!S4:S1000)+4)),0)").font = Font(italic=True, size=9)
        
        ws.cell(row=growth_rate_row, column=3, value=f"=IFERROR(B{growth_rate_row},0)").number_format = '+0.00;-0.00'
        ws.cell(row=growth_rate_row, column=4, value="percentage")
        
        ws.cell(row=15, column=1, value="PERFORMANCE SCORE CALCULATION (FORMULAS)").font = Font(bold=True, size=12, color=self.primary_color)
        
        score_headers = ["Component", "Weight", "Score", "Weighted Score"]
        for col, header in enumerate(score_headers, 1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        score_components = [
            ("Delivery Rate", "=IFERROR(Constants!B7,0)", f"=IFERROR(MIN(Constants!B13,(C{delivery_rate_row}*100)/Constants!B4),0)"),
            ("On-Time Rate", "=IFERROR(Constants!B8,0)", f"=IFERROR(MIN(Constants!B13,(C{on_time_rate_row}*100)/Constants!B5),0)"),
            ("Activity Level", "=IFERROR(Constants!B9,0)", f"=IFERROR(MIN(Constants!B13,C{total_shipments_row}/Constants!B6*100),0)"),
            ("Consistency", "=IFERROR(Constants!B10,0)", f"=IFERROR(MAX(0,Constants!B13-(C{cancel_rate_row}*Constants!B14)),0)"),
            ("Growth", "=IFERROR(Constants!B11,0)", f"=IFERROR(MAX(0,MIN(Constants!B13,Constants!B12+(C{growth_rate_row}*100))),0)")
        ]
        
        score_row = 18
        for component, weight_formula, score_formula in score_components:
            ws.cell(row=score_row, column=1, value=component).font = Font(bold=True)
            ws.cell(row=score_row, column=2, value=weight_formula).number_format = '0.0%'
            ws.cell(row=score_row, column=3, value=score_formula).number_format = '0.00'
            ws.cell(row=score_row, column=4, value=f"=IFERROR(B{score_row}*C{score_row},0)").number_format = '0.00'
            score_row += 1
        
        total_score_row = 23
        ws.cell(row=total_score_row, column=1, value="TOTAL PERFORMANCE SCORE").font = Font(bold=True, size=11)
        ws.cell(row=total_score_row, column=1).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        ws.cell(row=total_score_row, column=4, value=f"=IFERROR(SUM(D18:D22),0)").font = Font(bold=True, size=12, color=self.secondary_color)
        ws.cell(row=total_score_row, column=4).number_format = '0.00'
        
        ws.cell(row=25, column=1, value="NOTES ON CONSTANTS:").font = Font(bold=True, size=10, color=self.primary_color)
        ws.cell(row=26, column=1, value="All weight values and thresholds are defined in the hidden 'Constants' sheet").font = Font(size=9, italic=True)
        ws.cell(row=27, column=1, value="These are fixed system parameters, not calculated from data").font = Font(size=9, italic=True)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def create_cost_analysis_dashboard(self, wb, data, period_type):
        ws = wb.create_sheet("Cost Analysis")
        shipment_data = data.get('shipmentData', [])
        shipment_count = len(shipment_data)
        last_data_row = 3 + shipment_count
        
        title = ws.cell(row=1, column=1, value=f"COST ANALYSIS - {period_type.upper()} EXCEL FORMULAS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        ws.cell(row=2, column=1, value="All cost calculations use formulas referencing Shipment Data sheet").font = Font(size=9, italic=True, color="6B7280")
        ws.merge_cells("A2:F2")
        
        ws.cell(row=4, column=1, value="COST METRICS (FORMULAS)").font = Font(bold=True, size=12, color=self.primary_color)
        
        headers = ["Metric", "Formula", "Result", "Unit"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if shipment_count > 0:
            total_cost_row = 6
            ws.cell(row=total_cost_row, column=1, value="Total Cost").font = Font(bold=True)
            ws.cell(row=total_cost_row, column=2, value=f"=IFERROR(SUM('Shipment Data'!T4:T{last_data_row}),0)").font = Font(italic=True, size=9)
            ws.cell(row=total_cost_row, column=3, value=f"=IFERROR(B{total_cost_row},0)").number_format = 'Rp #,##0'
            ws.cell(row=total_cost_row, column=4, value="IDR")
            
            avg_cost_row = 7
            ws.cell(row=avg_cost_row, column=1, value="Average Cost per Delivery").font = Font(bold=True)
            ws.cell(row=avg_cost_row, column=2, value=f"=IFERROR(IF(COUNTA('Shipment Data'!A4:A{last_data_row})=0,0,C{total_cost_row}/COUNTA('Shipment Data'!A4:A{last_data_row})),0)").font = Font(italic=True, size=9)
            ws.cell(row=avg_cost_row, column=3, value=f"=IFERROR(B{avg_cost_row},0)").number_format = 'Rp #,##0'
            ws.cell(row=avg_cost_row, column=4, value="IDR/delivery")
            
            cost_per_km_row = 8
            ws.cell(row=cost_per_km_row, column=1, value="Cost per Kilometer").font = Font(bold=True)
            ws.cell(row=cost_per_km_row, column=2, value=f"=IFERROR(IF(SUM('Shipment Data'!U4:U{last_data_row})=0,0,C{total_cost_row}/SUM('Shipment Data'!U4:U{last_data_row})),0)").font = Font(italic=True, size=9)
            ws.cell(row=cost_per_km_row, column=3, value=f"=IFERROR(B{cost_per_km_row},0)").number_format = 'Rp #,##0'
            ws.cell(row=cost_per_km_row, column=4, value="IDR/km")
            
            total_distance_row = 9
            ws.cell(row=total_distance_row, column=1, value="Total Distance").font = Font(bold=True)
            ws.cell(row=total_distance_row, column=2, value=f"=IFERROR(SUM('Shipment Data'!U4:U{last_data_row}),0)").font = Font(italic=True, size=9)
            ws.cell(row=total_distance_row, column=3, value=f"=IFERROR(B{total_distance_row},0)").number_format = '#,##0.00'
            ws.cell(row=total_distance_row, column=4, value="km")
            
            total_deliveries_row = 10
            ws.cell(row=total_deliveries_row, column=1, value="Total Deliveries").font = Font(bold=True)
            ws.cell(row=total_deliveries_row, column=2, value=f"=IFERROR(COUNTA('Shipment Data'!A4:A{last_data_row}),0)").font = Font(italic=True, size=9)
            ws.cell(row=total_deliveries_row, column=3, value=f"=IFERROR(B{total_deliveries_row},0)").number_format = '#,##0'
            ws.cell(row=total_deliveries_row, column=4, value="deliveries")
        else:
            ws.cell(row=6, column=1, value="No shipment data available").font = Font(color="FF0000")
            return
        
        ws.cell(row=13, column=1, value="COST BY PROJECT (FORMULAS)").font = Font(bold=True, size=12, color=self.primary_color)
        
        project_headers = ["Project Name", "Total Cost", "Deliveries", "Avg Cost", "Share"]
        for col, header in enumerate(project_headers, 1):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        projects = {}
        for shipment in shipment_data:
            project = self.clean_string(shipment.get('project_name'))
            if project != '-':
                if project not in projects:
                    projects[project] = {'cost': 0, 'count': 0}
                projects[project]['cost'] += self.clean_number(shipment.get('cost'))
                projects[project]['count'] += 1
        
        sorted_projects = sorted(projects.items(), key=lambda x: x[1]['cost'], reverse=True)[:10]
        
        project_row = 15
        for project_name, project_data in sorted_projects:
            ws.cell(row=project_row, column=1, value=project_name).font = Font(bold=True, size=9)
            ws.cell(row=project_row, column=2, value=f"=IFERROR(SUMIF('Shipment Data'!B:B,A{project_row},'Shipment Data'!T:T),0)").number_format = 'Rp #,##0'
            ws.cell(row=project_row, column=3, value=f"=IFERROR(COUNTIF('Shipment Data'!B:B,A{project_row}),0)").number_format = '#,##0'
            ws.cell(row=project_row, column=4, value=f"=IFERROR(IF(C{project_row}=0,0,B{project_row}/C{project_row}),0)").number_format = 'Rp #,##0'
            ws.cell(row=project_row, column=5, value=f"=IFERROR(B{project_row}/C{total_cost_row},0)").number_format = '0.0%'
            project_row += 1
        
        ws.cell(row=project_row + 2, column=1, value="COST BY HUB (FORMULAS)").font = Font(bold=True, size=12, color=self.primary_color)
        
        hub_headers = ["Hub Name", "Total Cost", "Deliveries", "Avg Cost", "Share"]
        hub_header_row = project_row + 3
        for col, header in enumerate(hub_headers, 1):
            cell = ws.cell(row=hub_header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        hubs = {}
        for shipment in shipment_data:
            hub = self.clean_string(shipment.get('hub'))
            if hub != '-':
                if hub not in hubs:
                    hubs[hub] = {'cost': 0, 'count': 0}
                hubs[hub]['cost'] += self.clean_number(shipment.get('cost'))
                hubs[hub]['count'] += 1
        
        sorted_hubs = sorted(hubs.items(), key=lambda x: x[1]['cost'], reverse=True)[:10]
        
        hub_row = hub_header_row + 1
        for hub_name, hub_data in sorted_hubs:
            ws.cell(row=hub_row, column=1, value=hub_name).font = Font(bold=True, size=9)
            ws.cell(row=hub_row, column=2, value=f"=IFERROR(SUMIF('Shipment Data'!E:E,A{hub_row},'Shipment Data'!T:T),0)").number_format = 'Rp #,##0'
            ws.cell(row=hub_row, column=3, value=f"=IFERROR(COUNTIF('Shipment Data'!E:E,A{hub_row}),0)").number_format = '#,##0'
            ws.cell(row=hub_row, column=4, value=f"=IFERROR(IF(C{hub_row}=0,0,B{hub_row}/C{hub_row}),0)").number_format = 'Rp #,##0'
            ws.cell(row=hub_row, column=5, value=f"=IFERROR(B{hub_row}/C{total_cost_row},0)").number_format = '0.0%'
            hub_row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
    
    def create_trend_analysis_with_formulas(self, wb, data, period_type):
        ws = wb.create_sheet("Delivery Trends")
        shipment_data = data.get('shipmentData', [])
        shipment_count = len(shipment_data)
        last_data_row = 3 + shipment_count
        
        title = ws.cell(row=1, column=1, value=f"DELIVERY TRENDS - {period_type.upper()} FORMULAS (CHRONOLOGICAL)")
        title.font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells("A1:F1")
        
        ws.cell(row=2, column=1, value="All trends calculated using formulas from Shipment Data | Sorted: Oldest first").font = Font(size=9, italic=True, color="6B7280")
        ws.merge_cells("A2:F2")
        
        ws.cell(row=4, column=1, value=f"{period_type.upper()} TREND DATA (FORMULAS)").font = Font(bold=True, size=11, color=self.primary_color)
        
        trend_headers = ["Period", "Deliveries", "Cost", "Cumulative", "Growth", "Status"]
        
        for col, header in enumerate(trend_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        periods_dict = {}
        for shipment in shipment_data:
            display_period, month_num, year_num, sort_key = self.extract_period_info(shipment.get('delivery_date'), period_type)
            if display_period and sort_key:
                if display_period not in periods_dict:
                    periods_dict[display_period] = {
                        'cost': 0,
                        'count': 0,
                        'sort_key': sort_key,
                        'month_num': month_num,
                        'year_num': year_num
                    }
                periods_dict[display_period]['cost'] += self.clean_number(shipment.get('cost'))
                periods_dict[display_period]['count'] += 1
        
        sorted_periods = sorted(periods_dict.items(), key=lambda x: x[1]['sort_key'])
        
        row_num = 6
        for idx, (period_display, period_data) in enumerate(sorted_periods):
            ws.cell(row=row_num, column=1, value=period_display)
            
            if period_type == 'monthly':
                month_num = period_data['month_num']
                year_num = period_data['year_num']
                ws.cell(row=row_num, column=2, value=f"=IFERROR(SUMPRODUCT(('Shipment Data'!$V:$V={month_num})*('Shipment Data'!$W:$W={year_num})*1),0)").number_format = '#,##0'
                ws.cell(row=row_num, column=3, value=f"=IFERROR(SUMIFS('Shipment Data'!$T:$T,'Shipment Data'!$V:$V,{month_num},'Shipment Data'!$W:$W,{year_num}),0)").number_format = 'Rp #,##0'
            else:
                ws.cell(row=row_num, column=2, value=f"=IFERROR(COUNTIF('Shipment Data'!Q:Q,A{row_num}),0)").number_format = '#,##0'
                ws.cell(row=row_num, column=3, value=f"=IFERROR(SUMIF('Shipment Data'!Q:Q,A{row_num},'Shipment Data'!T:T),0)").number_format = 'Rp #,##0'
            
            ws.cell(row=row_num, column=4, value=f"=IFERROR(SUM($B$6:B{row_num}),0)").number_format = '#,##0'
            
            if idx > 0:
                ws.cell(row=row_num, column=5, value=f"=IFERROR(IF(B{row_num-1}=0,0,(B{row_num}-B{row_num-1})/B{row_num-1}),0)").number_format = '+0.0%;-0.0%;0.0%'
                ws.cell(row=row_num, column=6, value=f'=IF(E{row_num}>0.1,"↑ Increasing",IF(E{row_num}>0,"↗ Growth",IF(E{row_num}<-0.1,"↓ Decreasing","→ Stable")))')
            else:
                ws.cell(row=row_num, column=5, value=0).number_format = '0.0%'
                ws.cell(row=row_num, column=6, value="Baseline")
            
            row_num += 1
        
        project_start = row_num + 3
        ws.cell(row=project_start, column=1, value="TOP PROJECTS BY DELIVERY VOLUME (FORMULAS)").font = Font(bold=True, size=11, color=self.primary_color)
        
        project_headers = ["Rank", "Project Name", "Deliveries", "Total Cost", "Percentage"]
        for col, header in enumerate(project_headers, 1):
            cell = ws.cell(row=project_start + 1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        projects = {}
        for shipment in shipment_data:
            project = self.clean_string(shipment.get('project_name'))
            if project != '-':
                if project not in projects:
                    projects[project] = {'count': 0, 'cost': 0}
                projects[project]['count'] += 1
                projects[project]['cost'] += self.clean_number(shipment.get('cost'))
        
        sorted_projects = sorted(projects.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
        
        project_row = project_start + 2
        total_range_start = project_row
        for idx, (project_name, project_data) in enumerate(sorted_projects, 1):
            ws.cell(row=project_row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=project_row, column=2, value=project_name).font = Font(bold=True, size=9)
            ws.cell(row=project_row, column=3, value=f"=IFERROR(COUNTIF('Shipment Data'!B:B,B{project_row}),0)").number_format = '#,##0'
            ws.cell(row=project_row, column=4, value=f"=SUMIF('Shipment Data'!B:B,B{project_row},'Shipment Data'!T:T)").number_format = 'Rp #,##0'
            
            if idx <= 3:
                ws.cell(row=project_row, column=2).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            
            project_row += 1
        
        total_range_end = project_row - 1
        for row_idx in range(total_range_start, total_range_end + 1):
            ws.cell(row=row_idx, column=5, value=f"=IF(SUM($C${total_range_start}:$C${total_range_end})=0,0,C{row_idx}/SUM($C${total_range_start}:$C${total_range_end}))").number_format = '0.0%'
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_project_analysis_with_formulas(self, wb, data, period_type):
        ws = wb.create_sheet("Project Analysis")
        shipment_data = data.get('shipmentData', [])
        shipment_count = len(shipment_data)
        
        if shipment_count == 0:
            ws.cell(row=1, column=1, value="NO SHIPMENT DATA AVAILABLE").font = Font(bold=True, size=16, color="FF0000")
            return
        
        title = ws.cell(row=1, column=1, value=f"PROJECT ANALYSIS - {period_type.upper()} FORMULAS")
        title.font = Font(bold=True, size=16, color=self.primary_color)
        ws.merge_cells("A1:G1")
        
        ws.cell(row=2, column=1, value="All project metrics calculated using formulas from Shipment Data").font = Font(size=9, italic=True, color="6B7280")
        ws.merge_cells("A2:G2")
        
        headers = ["Project Name", "Total Deliveries", "Total Cost", "Avg Cost", "Avg Distance", "On-Time Count", "On-Time Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        projects = {}
        for shipment in shipment_data:
            project = self.clean_string(shipment.get('project_name'))
            if project != '-':
                if project not in projects:
                    projects[project] = {'count': 0, 'cost': 0, 'distance': 0, 'on_time': 0}
                projects[project]['count'] += 1
                projects[project]['cost'] += self.clean_number(shipment.get('cost'))
                projects[project]['distance'] += self.clean_number(shipment.get('distance_km'))
                projects[project]['on_time'] += self.is_on_time(shipment.get('sla'))
        
        sorted_projects = sorted(projects.items(), key=lambda x: x[1]['count'], reverse=True)[:20]
        
        project_row = 5
        for project_name, project_data in sorted_projects:
            ws.cell(row=project_row, column=1, value=project_name).font = Font(bold=True, size=9)
            ws.cell(row=project_row, column=2, value=f"=COUNTIF('Shipment Data'!B:B,A{project_row})").number_format = '#,##0'
            ws.cell(row=project_row, column=3, value=f"=SUMIF('Shipment Data'!B:B,A{project_row},'Shipment Data'!T:T)").number_format = 'Rp #,##0'
            ws.cell(row=project_row, column=4, value=f"=IF(B{project_row}=0,0,C{project_row}/B{project_row})").number_format = 'Rp #,##0'
            ws.cell(row=project_row, column=5, value=f"=IF(B{project_row}=0,0,SUMIF('Shipment Data'!B:B,A{project_row},'Shipment Data'!U:U)/B{project_row})").number_format = '0.00'
            ws.cell(row=project_row, column=6, value=f"=SUMIFS('Shipment Data'!P:P,'Shipment Data'!B:B,A{project_row})").number_format = '#,##0'
            ws.cell(row=project_row, column=7, value=f"=IF(B{project_row}=0,0,F{project_row}/B{project_row})").number_format = '0.0%'
            project_row += 1
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def create_performance_overview_sheet(self, wb, data, has_valid_trends, period_type):
        ws = wb.create_sheet("Performance Overview")
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value=f"📊 PERFORMANCE OVERVIEW - {period_type.upper()} FORMULA-DRIVEN")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        subtitle = ws.cell(row=2, column=1, value="All KPIs calculated using Excel formulas")
        subtitle.font = Font(size=11, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:P2")
        
        current_row = 4
        
        ws.cell(row=current_row, column=2, value="KEY PERFORMANCE METRICS (FORMULAS)").font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells(f"B{current_row}:H{current_row}")
        current_row += 2
        
        kpi_metrics = [
            ("Total Deliveries", "='Performance Metrics'!C4", "🚚", "E3F2FD", '#,##0'),
            ("Success Rate", "='Performance Metrics'!C10*100", "✅", "E8F5E9", '0.0"%"'),
            ("On-Time Rate", "='Performance Metrics'!C6*100", "⏰", "FFF3E0", '0.0"%"'),
            ("Total Cost", "='Cost Analysis'!C6", "💰", "FFF4E6", 'Rp #,##0'),
            ("Avg Cost/Delivery", "='Cost Analysis'!C7", "💵", "F3E5F5", 'Rp #,##0'),
            ("Cost per Km", "='Cost Analysis'!C8", "📏", "E1F5FE", 'Rp #,##0')
        ]
        
        kpi_row = current_row
        col_offset = 2
        
        for idx, (label, formula, icon, bg_color, num_format) in enumerate(kpi_metrics):
            if idx % 3 == 0 and idx > 0:
                kpi_row += 5
                col_offset = 2
            
            ws.cell(row=kpi_row, column=col_offset, value=icon).font = Font(size=18)
            ws.cell(row=kpi_row, column=col_offset).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=kpi_row, column=col_offset).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row}:{get_column_letter(col_offset+1)}{kpi_row}")
            
            ws.cell(row=kpi_row+1, column=col_offset, value=label).font = Font(bold=True, size=9)
            ws.cell(row=kpi_row+1, column=col_offset).alignment = Alignment(horizontal="center")
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row+1}:{get_column_letter(col_offset+1)}{kpi_row+1}")
            
            value_cell = ws.cell(row=kpi_row+2, column=col_offset, value=formula)
            value_cell.font = Font(bold=True, size=14, color=self.secondary_color)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.number_format = num_format
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row+2}:{get_column_letter(col_offset+1)}{kpi_row+2}")
            
            col_offset += 3
        
        current_row = kpi_row + 8
        
        footer_row = current_row + 2
        ws.cell(row=footer_row, column=2, value=f"Performance Overview: {datetime.now().strftime('%d %B %Y, %H:%M')} | {period_type.upper()} formulas").font = Font(size=9, italic=True, color="6B7280")
        
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 4
    
    def create_operational_insights_dashboard(self, wb, data, period_type):
        ws = wb.create_sheet("Operational Insights")
        shipment_data = data.get('shipmentData', [])
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value=f"⚙️ OPERATIONAL INSIGHTS - {period_type.upper()} FORMULAS")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        subtitle = ws.cell(row=2, column=1, value="Operational Metrics with Excel Formulas")
        subtitle.font = Font(size=11, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:P2")
        
        current_row = 4
        
        if len(shipment_data) > 0:
            ws.cell(row=current_row, column=1, value=f"📍 HUB COST PERFORMANCE ({period_type.upper()} FORMULAS)").font = Font(bold=True, size=13, color=self.secondary_color)
            current_row += 1
            
            hubs = {}
            for shipment in shipment_data:
                hub = self.clean_string(shipment.get('hub'))
                if hub != '-':
                    if hub not in hubs:
                        hubs[hub] = {'cost': 0, 'count': 0}
                    hubs[hub]['cost'] += self.clean_number(shipment.get('cost'))
                    hubs[hub]['count'] += 1
            
            sorted_hubs = sorted(hubs.items(), key=lambda x: x[1]['cost'], reverse=True)[:10]
            
            hub_ws = wb.create_sheet("Hub Cost Data")
            hub_ws.sheet_state = 'hidden'
            
            hub_ws.cell(row=1, column=1, value="Hub")
            hub_ws.cell(row=1, column=2, value="Total Cost")
            hub_ws.cell(row=1, column=3, value="Deliveries")
            
            for idx, (hub_name, hub_data) in enumerate(sorted_hubs, 2):
                hub_ws.cell(row=idx, column=1, value=hub_name)
                hub_ws.cell(row=idx, column=2, value=f"=SUMIF('Shipment Data'!E:E,'Hub Cost Data'!A{idx},'Shipment Data'!T:T)")
                hub_ws.cell(row=idx, column=3, value=f"=COUNTIF('Shipment Data'!E:E,'Hub Cost Data'!A{idx})")
            
            hub_chart = BarChart()
            hub_chart.title = f"Hub Cost Distribution ({period_type.capitalize()} Formula-Based)"
            hub_chart.type = "col"
            hub_chart.style = 11
            hub_chart.y_axis.title = "Total Cost (IDR)"
            hub_chart.x_axis.title = "Hub Name"
            hub_chart.height = 14
            hub_chart.width = 28
            
            hub_data_ref = Reference(hub_ws, min_col=2, min_row=1, max_row=1+len(sorted_hubs))
            hub_cats_ref = Reference(hub_ws, min_col=1, min_row=2, max_row=1+len(sorted_hubs))
            hub_chart.add_data(hub_data_ref, titles_from_data=True)
            hub_chart.set_categories(hub_cats_ref)
            
            series_hub = hub_chart.series[0]
            series_hub.graphicalProperties.solidFill = "EC4899"
            
            ws.add_chart(hub_chart, f"A{current_row}")
            current_row += 28
        
        footer_row = current_row + 2
        ws.cell(row=footer_row, column=1, value=f"Operational Insights: {datetime.now().strftime('%d %B %Y, %H:%M')} | {period_type.upper()} Formula-driven").font = Font(size=9, italic=True, color="6B7280")
        
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 3
    
    def create_visual_dashboard(self, wb, data, has_valid_trends, period_type):
        ws = wb.create_sheet("Visual Charts")
        shipment_data = data.get('shipmentData', [])
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value=f"📊 VISUAL CHARTS - {period_type.upper()} FORMULA-DRIVEN")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        subtitle_text = f"Interactive Charts with {period_type.capitalize()} Formula Calculations" if has_valid_trends else f"⚠️ Limited Charts - Trend Analysis Unavailable"
        subtitle = ws.cell(row=2, column=1, value=subtitle_text)
        subtitle.font = Font(size=11, color="6B7280" if has_valid_trends else self.warning_color, italic=True, bold=not has_valid_trends)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:P2")
        
        current_row = 4
        
        if has_valid_trends and len(shipment_data) > 1:
            periods_dict = {}
            for shipment in shipment_data:
                display_period, month_num, year_num, sort_key = self.extract_period_info(shipment.get('delivery_date'), period_type)
                if display_period and sort_key:
                    if display_period not in periods_dict:
                        periods_dict[display_period] = {
                            'cost': 0,
                            'count': 0,
                            'sort_key': sort_key,
                            'month_num': month_num,
                            'year_num': year_num
                        }
                    periods_dict[display_period]['cost'] += self.clean_number(shipment.get('cost'))
                    periods_dict[display_period]['count'] += 1
            
            if len(periods_dict) >= 2:
                ws.cell(row=current_row, column=1, value=f"📈 {period_type.upper()} DELIVERY & COST TREND (FORMULAS)").font = Font(bold=True, size=13, color=self.secondary_color)
                current_row += 1
                
                trend_ws = wb.create_sheet("Trend Chart Data")
                trend_ws.sheet_state = 'hidden'
                
                trend_ws.cell(row=1, column=1, value="Period")
                trend_ws.cell(row=1, column=2, value="Deliveries")
                trend_ws.cell(row=1, column=3, value="Cost")
                trend_ws.cell(row=1, column=4, value="Month Num")
                trend_ws.cell(row=1, column=5, value="Year Num")
                
                sorted_periods = sorted(periods_dict.items(), key=lambda x: x[1]['sort_key'])
                
                for idx, (period_display, period_data) in enumerate(sorted_periods, 2):
                    trend_ws.cell(row=idx, column=1, value=period_display)
                    
                    if period_type == 'monthly':
                        month_num = period_data['month_num']
                        year_num = period_data['year_num']
                        trend_ws.cell(row=idx, column=2, value=f"=IFERROR(SUMPRODUCT(('Shipment Data'!$V:$V={month_num})*('Shipment Data'!$W:$W={year_num})*1),0)")
                        trend_ws.cell(row=idx, column=3, value=f"=IFERROR(SUMIFS('Shipment Data'!$T:$T,'Shipment Data'!$V:$V,{month_num},'Shipment Data'!$W:$W,{year_num}),0)")
                    else:
                        trend_ws.cell(row=idx, column=2, value=f"=IFERROR(COUNTIF('Shipment Data'!Q:Q,A{idx}),0)")
                        trend_ws.cell(row=idx, column=3, value=f"=IFERROR(SUMIF('Shipment Data'!Q:Q,A{idx},'Shipment Data'!T:T),0)")
                    
                    trend_ws.cell(row=idx, column=4, value=period_data.get('month_num', ''))
                    trend_ws.cell(row=idx, column=5, value=period_data.get('year_num', ''))
                
                line_chart = LineChart()
                line_chart.title = f"{period_type.capitalize()} Delivery Volume (Formula-Based)"
                line_chart.style = 12
                line_chart.y_axis.title = "Number of Deliveries"
                line_chart.x_axis.title = "Period"
                line_chart.height = 12
                line_chart.width = 24
                
                data_ref = Reference(trend_ws, min_col=2, min_row=1, max_row=1+len(sorted_periods))
                cats_ref = Reference(trend_ws, min_col=1, min_row=2, max_row=1+len(sorted_periods))
                line_chart.add_data(data_ref, titles_from_data=True)
                line_chart.set_categories(cats_ref)
                
                series_line = line_chart.series[0]
                series_line.graphicalProperties.line.width = 35000
                series_line.graphicalProperties.line.solidFill = "3B82F6"
                
                ws.add_chart(line_chart, f"A{current_row}")
                
                cost_chart = LineChart()
                cost_chart.title = f"{period_type.capitalize()} Cost Trend (Formula-Based)"
                cost_chart.style = 12
                cost_chart.y_axis.title = "Total Cost (IDR)"
                cost_chart.x_axis.title = "Period"
                cost_chart.height = 12
                cost_chart.width = 24
                
                cost_data_ref = Reference(trend_ws, min_col=3, min_row=1, max_row=1+len(sorted_periods))
                cost_chart.add_data(cost_data_ref, titles_from_data=True)
                cost_chart.set_categories(cats_ref)
                
                series_cost = cost_chart.series[0]
                series_cost.graphicalProperties.line.width = 35000
                series_cost.graphicalProperties.line.solidFill = "F59E0B"
                
                ws.add_chart(cost_chart, f"P{current_row}")
                current_row += 24
        else:
            ws.cell(row=current_row, column=1, value="⚠️ TREND CHARTS UNAVAILABLE").font = Font(bold=True, size=13, color=self.warning_color)
            current_row += 1
            periods_dict = {}
            for shipment in shipment_data:
                display_period, _, _, sort_key = self.extract_period_info(shipment.get('delivery_date'), period_type)
                if display_period and sort_key:
                    periods_dict[display_period] = sort_key
            period_count = len(periods_dict)
            ws.cell(row=current_row, column=1, value=f"Requires 2+ delivery periods. Current: {period_count} period(s)").font = Font(size=10, italic=True)
            ws.merge_cells(f"A{current_row}:P{current_row}")
            current_row += 3
        
        footer_row = current_row + 2
        analysis_type = f"{period_type.capitalize()} Formula-Based Analysis" if has_valid_trends else f"Limited {period_type.capitalize()} Analysis"
        ws.cell(row=footer_row, column=1, value=f"Charts Generated: {datetime.now().strftime('%d %B %Y, %H:%M')} | Type: {analysis_type}").font = Font(size=9, italic=True, color="6B7280")
        
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 3
    
    def create_advanced_analytics_dashboard(self, wb, data, period_type):
        ws = wb.create_sheet("Advanced Analytics")
        shipment_data = data.get('shipmentData', [])
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value=f"📊 ADVANCED ANALYTICS - {period_type.upper()} FORMULAS")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        subtitle = ws.cell(row=2, column=1, value="Deep Dive Analysis with Excel Formulas")
        subtitle.font = Font(size=11, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:P2")
        
        current_row = 4
        
        periods_dict = {}
        for shipment in shipment_data:
            display_period, month_num, year_num, sort_key = self.extract_period_info(shipment.get('delivery_date'), period_type)
            if display_period and sort_key:
                if display_period not in periods_dict:
                    periods_dict[display_period] = {
                        'cost': 0,
                        'count': 0,
                        'sort_key': sort_key,
                        'month_num': month_num,
                        'year_num': year_num
                    }
                periods_dict[display_period]['cost'] += self.clean_number(shipment.get('cost'))
                periods_dict[display_period]['count'] += 1
        
        if len(periods_dict) > 1:
            ws.cell(row=current_row, column=1, value=f"📈 COST EFFICIENCY TREND ({period_type.upper()} FORMULAS)").font = Font(bold=True, size=13, color=self.secondary_color)
            current_row += 1
            
            efficiency_ws = wb.create_sheet("Efficiency Data")
            efficiency_ws.sheet_state = 'hidden'
            
            efficiency_ws.cell(row=1, column=1, value="Period")
            efficiency_ws.cell(row=1, column=2, value="Cost per Delivery")
            
            sorted_periods = sorted(periods_dict.items(), key=lambda x: x[1]['sort_key'])
            
            for idx, (period_display, period_data) in enumerate(sorted_periods, 2):
                efficiency_ws.cell(row=idx, column=1, value=period_display)
                
                if period_type == 'monthly':
                    month_num = period_data['month_num']
                    year_num = period_data['year_num']
                    efficiency_ws.cell(row=idx, column=2, value=f"=IF(COUNTIF('Shipment Data'!V:V,{month_num})=0,0,SUMIFS('Shipment Data'!T:T,'Shipment Data'!V:V,{month_num})/COUNTIF('Shipment Data'!V:V,{month_num}))")
                else:
                    efficiency_ws.cell(row=idx, column=2, value=f"=IF(COUNTIF('Shipment Data'!Q:Q,A{idx})=0,0,SUMIF('Shipment Data'!Q:Q,A{idx},'Shipment Data'!T:T)/COUNTIF('Shipment Data'!Q:Q,A{idx})")
            
            efficiency_chart = LineChart()
            efficiency_chart.title = f"Cost Efficiency Trend ({period_type.capitalize()} Formula-Calculated)"
            efficiency_chart.style = 12
            efficiency_chart.y_axis.title = "Cost per Delivery (IDR)"
            efficiency_chart.x_axis.title = "Period"
            efficiency_chart.height = 14
            efficiency_chart.width = 28
            
            eff_data_ref = Reference(efficiency_ws, min_col=2, min_row=1, max_row=1+len(sorted_periods))
            eff_cats_ref = Reference(efficiency_ws, min_col=1, min_row=2, max_row=1+len(sorted_periods))
            efficiency_chart.add_data(eff_data_ref, titles_from_data=True)
            efficiency_chart.set_categories(eff_cats_ref)
            
            series_eff = efficiency_chart.series[0]
            series_eff.graphicalProperties.line.solidFill = "10B981"
            
            ws.add_chart(efficiency_chart, f"A{current_row}")
            current_row += 28
        
        footer_row = current_row + 2
        ws.cell(row=footer_row, column=1, value=f"Advanced Analytics: {datetime.now().strftime('%d %B %Y, %H:%M')} | {period_type.upper()} formulas").font = Font(size=9, italic=True, color="6B7280")
        
        for col in range(1, 33):
            ws.column_dimensions[get_column_letter(col)].width = 3
    
    def create_management_kpi_dashboard(self, wb, data, has_valid_trends, period_type):
        ws = wb.create_sheet("Management KPI")
        
        ws.sheet_view.showGridLines = False
        
        title = ws.cell(row=1, column=1, value=f"📊 MANAGEMENT KPI - {period_type.upper()} FORMULAS")
        title.font = Font(bold=True, size=18, color=self.primary_color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 30
        
        subtitle = ws.cell(row=2, column=1, value="Executive Summary with Formula-Based Calculations")
        subtitle.font = Font(size=11, color="6B7280", italic=True)
        subtitle.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:P2")
        
        current_row = 4
        
        ws.cell(row=current_row, column=2, value="KEY PERFORMANCE INDICATORS (FORMULAS)").font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells(f"B{current_row}:H{current_row}")
        current_row += 2
        
        kpi_data = [
            ("Total Deliveries", "='Performance Metrics'!C4", "🚚", "E3F2FD", "Volume", '#,##0'),
            ("Success Rate", "='Performance Metrics'!C10*100", "✅", "E8F5E9", "Quality", '0.0"%"'),
            ("On-Time Rate", "='Performance Metrics'!C6*100", "⏰", "FFF3E0", "Efficiency", '0.0"%"'),
            ("Total Cost", "='Cost Analysis'!C6", "💰", "FFF4E6", "Financial", 'Rp #,##0'),
            ("Avg Cost/Delivery", "='Cost Analysis'!C7", "💵", "F3E5F5", "Cost Efficiency", 'Rp #,##0'),
            ("Cost per Km", "='Cost Analysis'!C8", "📏", "E1F5FE", "Distance Cost", 'Rp #,##0')
        ]
        
        kpi_row = current_row
        col_offset = 2
        
        for idx, (label, formula, icon, bg_color, category, num_format) in enumerate(kpi_data):
            if idx % 3 == 0 and idx > 0:
                kpi_row += 6
                col_offset = 2
            
            ws.cell(row=kpi_row, column=col_offset, value=icon).font = Font(size=20)
            ws.cell(row=kpi_row, column=col_offset).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=kpi_row, column=col_offset).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row}:{get_column_letter(col_offset+1)}{kpi_row}")
            
            ws.cell(row=kpi_row+1, column=col_offset, value=label).font = Font(bold=True, size=10)
            ws.cell(row=kpi_row+1, column=col_offset).alignment = Alignment(horizontal="center")
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row+1}:{get_column_letter(col_offset+1)}{kpi_row+1}")
            
            value_cell = ws.cell(row=kpi_row+2, column=col_offset, value=formula)
            value_cell.font = Font(bold=True, size=16, color=self.secondary_color)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.number_format = num_format
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row+2}:{get_column_letter(col_offset+1)}{kpi_row+2}")
            
            ws.cell(row=kpi_row+3, column=col_offset, value=category).font = Font(size=8, italic=True,color="6B7280")
            ws.merge_cells(f"{get_column_letter(col_offset)}{kpi_row+3}:{get_column_letter(col_offset+1)}{kpi_row+3}")
            
            col_offset += 3
        
        current_row = kpi_row + 8
        
        ws.cell(row=current_row, column=2, value="STRATEGIC INSIGHTS").font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells(f"B{current_row}:H{current_row}")
        current_row += 2
        
        insights = [
            "• Performance score above 80 indicates operational excellence",
            "• On-time rate below 90% requires immediate attention",
            "• Cost per delivery above industry average needs optimization",
            "• Growth trends inform strategic expansion decisions",
            "• Consistent performance across periods indicates stability"
        ]
        
        for insight in insights:
            ws.cell(row=current_row, column=2, value=insight).font = Font(size=10)
            ws.merge_cells(f"B{current_row}:H{current_row}")
            current_row += 1
        
        footer_row = current_row + 2
        ws.cell(row=footer_row, column=2, value=f"Management KPI Dashboard: {datetime.now().strftime('%d %B %Y, %H:%M')} | {period_type.upper()} formulas").font = Font(size=9, italic=True, color="6B7280")
        
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 4

def main():
    try:
        if len(sys.argv) != 3:
            raise ValueError("Usage: python mitraPerformanceChartGeneratorFormula.py <input_json> <output_excel>")
        
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        shipment_data = data.get('shipmentData', [])
        period_type = data.get('periodType', 'monthly')
        
        if len(shipment_data) == 0:
            print(json.dumps({
                "success": False,
                "error": "No shipment data available. Cannot generate report without shipment data.",
                "details": "Please ensure delivery data is available before generating the report."
            }))
            sys.exit(1)
        
        generator = MitraPerformanceChartGeneratorFormula()
        result_path = generator.create_workbook_with_charts(data, output_path)
        
        data_quality = data.get('dataQuality', {})
        has_valid_trends = data_quality.get('hasValidTrends', False)
        trend_count = data_quality.get('trendCount', 0)
        
        message = f"Mitra performance chart with {period_type} Excel formulas created successfully"
        if not has_valid_trends:
            message = f"Limited analysis report created (only {trend_count} period available). Add more delivery periods for full features."
        
        print(json.dumps({
            "success": True,
            "output_path": result_path,
            "message": message,
            "data_quality": {
                "has_valid_trends": has_valid_trends,
                "trend_count": trend_count,
                "shipment_count": len(shipment_data),
                "period_type": period_type
            },
            "formula_info": {
                "all_calculations_use_formulas": True,
                "constants_sheet": "Constants (hidden)",
                "source_data_sheet": "Shipment Data",
                "period_filter": period_type,
                "note": f"All metrics are calculated dynamically using Excel formulas for transparency and auditability with {period_type} period filtering"
            }
        }))
    
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()